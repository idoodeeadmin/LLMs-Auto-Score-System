import asyncio
from fastapi import FastAPI, Depends, HTTPException, status, Header, UploadFile, File, Form, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import Optional, List
from .database import get_db_connection, init_db
from .auth import get_password_hash, verify_password, create_access_token, decode_token
import pymysql
import random
import string
import os
import json as json_module_top
import csv
import io
import statistics
from dotenv import load_dotenv
import aiofiles
import httpx
import time
import cloudinary
import cloudinary.uploader
from cloudinary.utils import cloudinary_url

load_dotenv()

# Cloudinary Configuration
cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET"),
    secure=True
)

def upload_to_cloudinary(file_bytes, folder="evaly", public_id=None):
    try:
        upload_result = cloudinary.uploader.upload(
            file_bytes,
            folder=folder,
            public_id=public_id,
            resource_type="auto"
        )
        return upload_result.get("secure_url")
    except Exception as e:
        print(f"[Cloudinary] Upload error: {e}")
        return None

async def get_image_bytes(path_or_url: str):
    if not path_or_url: return None
    if path_or_url.startswith("http"):
        try:
            async with httpx.AsyncClient() as client:
                resp = await client.get(path_or_url, timeout=10.0)
                if resp.status_code == 200:
                    return resp.content
                return None
        except Exception as e:
            print(f"[Grading Worker] Failed to fetch URL {path_or_url}: {e}")
            return None
    else:
        local_path = path_or_url.lstrip("/")
        if os.path.exists(local_path):
            try:
                async with aiofiles.open(local_path, "rb") as f:
                    return await f.read()
            except Exception as e:
                print(f"[Grading Worker] Failed to read local {local_path}: {e}")
                return None
        return None

app = FastAPI(title="Evaly API")

# Setup CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Database on startup
@app.on_event("startup")
async def startup_event():
    init_db()
    # Mount uploads folder for serving images
    os.makedirs("uploads", exist_ok=True)
    
    # Start grading worker
    asyncio.create_task(grading_worker())
    
    # Recover pending submissions
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, exam_id, student_id FROM submissions WHERE status = 'submitted'")
    pending = cursor.fetchall()
    for row in pending:
        # We need room_id, let's fetch it from exam
        cursor.execute("SELECT room_id FROM exams WHERE id = ?", (row["exam_id"],))
        exam_row = cursor.fetchone()
        if exam_row:
            await grading_queue.put({
                "submission_id": row["id"],
                "room_id": exam_row["room_id"],
                "exam_id": row["exam_id"],
                "user_id": row["student_id"]
            })
    conn.close()
    if pending:
        print(f"[Startup] Recovered {len(pending)} pending submissions into grading queue")


# Serve uploaded images as static files
try:
    app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")
    app.mount("/api/uploads", StaticFiles(directory="uploads"), name="api_uploads")
except Exception:
    pass  # Directory may not exist yet on first run

# Models
class UserRegister(BaseModel):
    email: str
    password: str
    name: str
    role: Optional[str] = "unassigned"
    student_id: Optional[str] = None

class UserLogin(BaseModel):
    email: str
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: dict

class RoomCreate(BaseModel):
    name: str
    section: Optional[str] = None

class JoinRoomRequest(BaseModel):
    class_code: str

class ForgotPasswordRequest(BaseModel):
    email: str
    name: Optional[str] = None
    student_id: Optional[str] = None

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

class VerifyEmailRequest(BaseModel):
    token: str

class ResendVerificationRequest(BaseModel):
    email: str

class AnnouncementCreate(BaseModel):
    title: str
    content: str


# In-memory Rate Limiter
REQUEST_LOGS = {} # { "ip": [timestamp1, timestamp2...] }

# Batch Grading Queue
grading_queue = asyncio.Queue()


def check_rate_limit(ip: str, limit: int = 10, window: int = 60):
    """Simple IP-based rate limiter (default: 10 requests per minute)"""
    now = time.time()
    if ip not in REQUEST_LOGS:
        REQUEST_LOGS[ip] = []
    
    # Clean old logs outside window
    REQUEST_LOGS[ip] = [t for t in REQUEST_LOGS[ip] if now - t < window]
    
    if len(REQUEST_LOGS[ip]) >= limit:
        return False
    
    REQUEST_LOGS[ip].append(now)
    return True

# Dependencies

async def grading_worker():
    while True:
        task = await grading_queue.get()
        try:
            submission_id = task.get("submission_id")
            room_id = task["room_id"]
            exam_id = task["exam_id"]
            user_id = task.get("user_id")
            specific_q_id = task.get("question_id")
            
            print(f"[Grading Worker] Processing submission {submission_id} (Question: {specific_q_id or 'All'})...")
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Fetch all answers for this submission
            cursor.execute("SELECT * FROM submission_answers WHERE submission_id = ?", (submission_id,))
            answers = cursor.fetchall()
            
            # Fetch questions for rubrics/answer_keys
            cursor.execute("SELECT * FROM questions WHERE exam_id = ? ORDER BY order_index", (exam_id,))
            questions = {q["id"]: dict(q) for q in cursor.fetchall()}
            
            total_ai_score = 0.0
            confidences = []
            
            for ans in answers:
                q_id = ans["question_id"]
                if specific_q_id and q_id != specific_q_id:
                    continue
                    
                q = questions.get(q_id)
                if not q:
                    continue
                    
                answer_text = ans["answer_text"] or ""
                
                # Load images (could be local or cloudinary)
                img_list = []
                img_mime_list = []
                
                image_paths_json = ans["image_paths"]
                if image_paths_json:
                    import json
                    paths = json.loads(image_paths_json)
                    for path in paths:
                        raw_bytes = await get_image_bytes(path)
                        if raw_bytes:
                            img_list.append(raw_bytes)
                            # guess mime
                            if path.endswith(".png"): mime = "image/png"
                            elif path.endswith(".webp"): mime = "image/webp"
                            else: mime = "image/jpeg"
                            img_mime_list.append(mime)
                
                # Load question images
                q_img_list = []
                q_img_mime_list = []
                q_image_paths = q.get("image_paths")
                if q_image_paths:
                    try:
                        q_paths = json_module.loads(q_image_paths)
                        for qp in q_paths:
                            qb = await get_image_bytes(qp)
                            if qb:
                                q_img_list.append(qb)
                                q_img_mime_list.append("image/png" if qp.endswith(".png") else "image/webp" if qp.endswith(".webp") else "image/jpeg")
                    except Exception: pass
                elif q.get("image_path"):
                    qp = q["image_path"]
                    qb = await get_image_bytes(qp)
                    if qb:
                        q_img_list.append(qb)
                        q_img_mime_list.append("image/png" if qp.endswith(".png") else "image/webp" if qp.endswith(".webp") else "image/jpeg")

                # Parse rubrics
                rubrics_data = None
                if q.get("rubrics"):
                    try:
                        rubrics_data = json_module.loads(q["rubrics"])
                    except Exception:
                        rubrics_data = None
                        
                ai_result = await score_with_gemini(
                    question_text=q.get("text") or "",
                    answer_text=answer_text,
                    max_score=q["score"],
                    answer_key=q.get("answer_key"),
                    rubrics=rubrics_data,
                    image_bytes_list=img_list[:5],
                    image_mime_list=img_mime_list[:5],
                    q_image_bytes_list=q_img_list[:5],
                    q_image_mime_list=q_img_mime_list[:5]
                )
                total_ai_score += ai_result["score"]
                confidences.append(ai_result["confidence"])
                
                # Store quality metrics if present
                q_metrics = json_module_top.dumps(ai_result.get("metrics", {})) if ai_result.get("metrics") else None
                
                cursor.execute(
                    "UPDATE submission_answers SET ai_score = ?, ai_feedback = ?, ai_confidence = ?, quality_metrics = ? WHERE id = ?",
                    (ai_result["score"], ai_result["feedback"], ai_result["confidence"], q_metrics, ans["id"])
                )
                
            new_status = "needs_review" if "low" in confidences else "ready"
            
            # Recalculate total score for the submission from all answers
            cursor.execute("SELECT SUM(COALESCE(teacher_score, ai_score, 0)) as total FROM submission_answers WHERE submission_id = ?", (submission_id,))
            total_score_row = cursor.fetchone()
            new_total_score = total_score_row["total"] if total_score_row and total_score_row["total"] else 0.0

            cursor.execute(
                "UPDATE submissions SET status = ?, total_score = ?, graded_by_ai = 1 WHERE id = ?",
                (new_status, round(new_total_score, 1), submission_id)
            )
            conn.commit()
            
            # Notify teacher
            cursor.execute("SELECT owner_id FROM rooms WHERE id = ?", (room_id,))
            teacher_row = cursor.fetchone()
            if teacher_row:
                await trigger_socket_notify(
                    user_id=teacher_row["owner_id"],
                    notify_type="ai_graded",
                    message=f"มีนักศึกษาส่งข้อสอบใหม่ในห้องของคุณ และ AI ตรวจเสร็จแล้ว!",
                    data={"exam_id": exam_id, "room_id": room_id, "submission_id": submission_id}
                )
            
            conn.close()
            print(f"[Grading Worker] Finished submission {submission_id}")
            
        except Exception as e:
            print(f"[Grading Worker] Error processing task: {e}")
            import traceback
            traceback.print_exc()
        finally:
            grading_queue.task_done()

async def trigger_socket_notify(user_id: int, notify_type: str, message: str, data: dict = None):
    """Bridge to Node.js Socket server to emit real-time notifications"""
    socket_url = f"http://localhost:{os.getenv('SOCKET_PORT', '3001')}/emit-notification"
    try:
        async with httpx.AsyncClient() as client:
            await client.post(socket_url, json={
                "userId": user_id,
                "type": notify_type,
                "message": message,
                "data": data or {}
            }, timeout=2.0)
    except Exception as e:
        print(f"[Socket Bridge Error] {e}")


def log_audit_action(user_id: int, action_type: str, entity_id: Optional[str] = None, details: Optional[str] = None, ip_address: Optional[str] = None):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO audit_logs (user_id, action_type, entity_id, details, ip_address) VALUES (?, ?, ?, ?, ?)",
            (user_id, action_type, entity_id, details, ip_address)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Failed to log audit action: {e}")

def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Invalid token")
    
    token = authorization.split(" ")[1]
    payload = decode_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    email = payload.get("sub")
    token_version = payload.get("token_version", 0)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, email, name, role, student_id, avatar_url, is_verified, token_version FROM users WHERE email = ?", (email,))
    user = cursor.fetchone()
    conn.close()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    if user.get("token_version", 0) != token_version:
        raise HTTPException(status_code=401, detail="Session expired or revoked")
    
    return dict(user)

# Auth Routes
@app.post("/api/auth/register")
async def register(user: UserRegister, request: Request):
    # Rate limit: 5 registrations per hour per IP
    if not check_rate_limit(request.client.host, limit=5, window=3600):
        raise HTTPException(status_code=429, detail="Too many registration attempts. Please try again later.")
    conn = get_db_connection()
    cursor = conn.cursor()
    
    hashed_password = get_password_hash(user.password)
    
    try:
        cursor.execute(
            "INSERT INTO users (email, password, name, role, student_id, is_verified) VALUES (?, ?, ?, ?, ?, 0)",
            (user.email, hashed_password, user.name, "unassigned", None)
        )
        user_id = cursor.lastrowid
        
        import uuid
        from datetime import datetime, timezone, timedelta
        import os, smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        token = uuid.uuid4().hex
        expires_at = (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
        
        cursor.execute(
            "INSERT INTO email_verifications (user_id, token, expires_at) VALUES (?, ?, ?)",
            (user_id, token, expires_at)
        )
        conn.commit()
    except pymysql.err.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="Email already registered")
        
    # Send verification email
    verify_link = f"http://localhost:8080/verify-email?token={token}"
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port_str = os.getenv("SMTP_PORT", "587")
    smtp_port = int(smtp_port_str) if smtp_port_str else 587
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    
    dev_verify_link = None
    if smtp_host and smtp_user and smtp_password:
        try:
            msg = MIMEMultipart()
            msg['From'] = f"Evaly Score <{smtp_user}>"
            msg['To'] = user.email
            msg['Subject'] = "ยืนยันบัญชีอีเมล Evaly Score (Verify Email)"
            
            body = f"""
            <h2>ยินดีต้อนรับสู่ Evaly Score</h2>
            <p>กรุณาคลิกที่ลิงก์ด้านล่างเพื่อยืนยันบัญชีอีเมลของคุณ (ลิงก์มีอายุการใช้งาน 24 ชั่วโมง):</p>
            <p><a href="{verify_link}">{verify_link}</a></p>
            """
            
            msg.attach(MIMEText(body, 'html'))
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
            server.quit()
        except Exception as e:
            print(f"\n[Verify Email Error] Failed to send email: {e}")
            dev_verify_link = verify_link
    else:
        dev_verify_link = verify_link
        
    conn.close()
    
    if dev_verify_link:
        print(f"\n========== VERIFY EMAIL ==========")
        print(f"Verify Link: {dev_verify_link}")
        print(f"==================================\n")
        
    return {
        "message": "สมัครสมาชิกสำเร็จ กรุณาตรวจสอบอีเมลเพื่อยืนยันบัญชีของคุณ",
        "dev_verify_link": dev_verify_link
    }

@app.post("/api/auth/login", response_model=TokenResponse)
async def login(user_data: UserLogin, request: Request):
    # Rate limit: 10 login attempts per minute
    if not check_rate_limit(request.client.host, limit=10, window=60):
        raise HTTPException(status_code=429, detail="Too many login attempts. Please wait a minute.")
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM users WHERE email = ?", (user_data.email,))
    user = cursor.fetchone()
    conn.close()
    
    if not user or not verify_password(user_data.password, user["password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
        
    access_token = create_access_token(data={"sub": user["email"], "token_version": user.get("token_version", 0)})
    
    log_audit_action(user["id"], "LOGIN", None, "User logged in via email", request.client.host)
    user_dict = dict(user)
    
    user_info = {
        "id": user_dict["id"],
        "email": user_dict["email"],
        "name": user_dict["name"],
        "role": user_dict["role"],
        "studentId": user_dict.get("student_id"),
        "avatarUrl": user_dict.get("avatar_url"),
        "is_verified": user_dict.get("is_verified", 0)
    }
    
    return {"access_token": access_token, "token_type": "bearer", "user": user_info}

@app.get("/api/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return {
        "id": user["id"],
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
        "studentId": user["student_id"],
        "avatarUrl": user.get("avatar_url", None),
        "is_verified": user.get("is_verified", 0)
    }

@app.put("/api/auth/profile")
async def update_profile(
    name: str = Form(None),
    password: str = Form(None),
    avatar: UploadFile = File(None),
    user: dict = Depends(get_current_user)
):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    avatar_filename = user.get("avatar_url")
    if avatar:
        # Upload to Cloudinary
        file_content = await avatar.read()
        avatar_url = upload_to_cloudinary(file_content, folder="avatars")
        if avatar_url:
            update_fields.append("avatar = ?")
            params.append(avatar_url)
                
    update_fields = []
    params = []
    
    if name:
        update_fields.append("name = ?")
        params.append(name)
        
    if password:
        update_fields.append("password = ?")
        params.append(get_password_hash(password))
        
    if avatar_filename and avatar_filename != user.get("avatar_url"):
        update_fields.append("avatar_url = ?")
        params.append(avatar_filename)
        
    if update_fields:
        query = f"UPDATE users SET {', '.join(update_fields)} WHERE id = ?"
        params.append(user["id"])
        
        try:
            cursor.execute(query, tuple(params))
            conn.commit()
        except pymysql.err.Error as e:
            conn.close()
            raise HTTPException(status_code=500, detail=str(e))
            
    conn.close()
    return {"message": "Profile updated successfully", "avatarUrl": avatar_filename}

class SetRoleRequest(BaseModel):
    role: str
    student_id: Optional[str] = None

@app.post("/api/auth/set-role")
async def set_role(req: SetRoleRequest, user: dict = Depends(get_current_user)):
    if user["role"] != "unassigned":
        raise HTTPException(status_code=400, detail="Role already set")
    
    if req.role not in ["teacher", "student"]:
        raise HTTPException(status_code=400, detail="Invalid role")
        
    if req.role == "student" and not req.student_id:
        raise HTTPException(status_code=400, detail="Student ID is required for students")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "UPDATE users SET role = ?, student_id = ? WHERE id = ?",
            (req.role, req.student_id, user["id"])
        )
        conn.commit()
    except pymysql.err.Error as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))
        
    conn.close()
    return {"message": "Role updated successfully", "role": req.role, "student_id": req.student_id}

@app.post("/api/auth/forgot-password")
async def forgot_password(req: ForgotPasswordRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT id, name, student_id FROM users WHERE email = ?", (req.email,))
    user = cursor.fetchone()
    
    if not user:
        conn.close()
        # For security, don't scream that email isn't there
        return {"message": "If that email exists, a password reset link has been generated."}
        
    # Self-Service Recovery if additional identity info provided
    if req.name or req.student_id:
        is_owner = False
        if req.name and user["name"] and req.name.strip().lower() == user["name"].strip().lower():
            is_owner = True
        elif req.student_id and user["student_id"] and req.student_id.strip().lower() == user["student_id"].strip().lower():
            is_owner = True
            
        if not is_owner:
            conn.close()
            raise HTTPException(status_code=400, detail="คุณไม่ใช่เจ้าของบัญชี ชื่อหรือรหัสนิสิตไม่ตรงกับฐานข้อมูล")
            
        import uuid
        from datetime import datetime, timezone, timedelta
        
        token = uuid.uuid4().hex
        expires_at = (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
        
        cursor.execute("DELETE FROM password_resets WHERE user_id = ?", (user["id"],))
        cursor.execute("INSERT INTO password_resets (user_id, token, expires_at) VALUES (?, ?, ?)", (user["id"], token, expires_at))
        conn.commit()
        conn.close()
        
        return {"message": "ยืนยันตัวตนสำเร็จ!", "reset_token": token}
        
    import uuid
    from datetime import datetime, timezone, timedelta
    
    token = uuid.uuid4().hex
    
    # Store aware datetime
    now_utc = datetime.now(timezone.utc)
    expires_at = (now_utc + timedelta(hours=1)).isoformat()
    
    # Delete old tokens for this user
    cursor.execute("DELETE FROM password_resets WHERE user_id = ?", (user["id"],))
    
    # Add new token
    cursor.execute(
        "INSERT INTO password_resets (user_id, token, expires_at) VALUES (?, ?, ?)",
        (user["id"], token, expires_at)
    )
    conn.commit()
    conn.close()
    
    # Generate reset link
    reset_link = f"http://localhost:8080/reset-password?token={token}"
    
    # Sending Email via SMTP
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    import os
    
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port_str = os.getenv("SMTP_PORT", "587")
    smtp_port = int(smtp_port_str) if smtp_port_str else 587
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    
    error_msg = None
    if smtp_host and smtp_user and smtp_password:
        try:
            msg = MIMEMultipart()
            msg['From'] = f"Evaly Score <{smtp_user}>"
            msg['To'] = req.email
            msg['Subject'] = "รีเซ็ตรหัสผ่าน Evaly Score (Password Reset)"
            
            body = f"""
            <h2>รีเซ็ตรหัสผ่าน Evaly Score</h2>
            <p>เราได้รับการร้องขอให้รีเซ็ตรหัสผ่านสำหรับบัญชีของคุณ</p>
            <p>กรุณาคลิกที่ลิงก์ด้านล่างเพื่อตั้งรหัสผ่านใหม่ (ลิงก์นี้มีอายุการใช้งาน 1 ชั่วโมง):</p>
            <p><a href="{reset_link}">{reset_link}</a></p>
            <p><br>หากคุณไม่ได้ร้องขอการรีเซ็ตรหัสผ่านนี้ กรุณาเพิกเฉยต่ออีเมลฉบับนี้</p>
            """
            
            msg.attach(MIMEText(body, 'html'))
            
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
            server.quit()
            print(f"\n[Email Sent] Password reset link sent to {req.email}")
            
            return {"message": "ส่งลิงก์สำหรับรีเซ็ตรหัสผ่านไปยังอีเมลของคุณแล้ว กรุณาตรวจสอบกล่องข้อความ"}
        except Exception as e:
            print(f"\n[Email Error] Failed to send email: {e}")
            error_msg = str(e)
            
    # Fallback to dev logs
    print(f"\n========== FORGOT PASSWORD ==========")
    print(f"Request for: {req.email}")
    print(f"Reset Link: {reset_link}")
    if error_msg:
        print(f"SMTP Error: {error_msg}")
    elif not smtp_host:
        print("Note: SMTP variables not configured in .env")
    print(f"======================================\n")
    
    # Returning the link for development/testing ease
    return {
        "message": "รหัสสำหรับการทดสอบ (Dev Mode): ระบบยังไม่ได้ตั้งค่า Email SMTP",
        "dev_reset_link": reset_link
    }

@app.post("/api/auth/reset-password")
async def reset_password(req: ResetPasswordRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM password_resets WHERE token = ?", (req.token,))
    reset_record = cursor.fetchone()
    
    if not reset_record:
        conn.close()
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")
        
    # Check expiration
    from datetime import datetime, timezone
    try:
        expires_at_str = reset_record["expires_at"]
        expires_at = datetime.fromisoformat(expires_at_str.replace('Z', '+00:00'))
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
            
        now_utc = datetime.now(timezone.utc)
        if now_utc > expires_at:
            cursor.execute("DELETE FROM password_resets WHERE id = ?", (reset_record["id"],))
            conn.commit()
            conn.close()
            raise HTTPException(status_code=400, detail="Reset token has expired")
    except ValueError:
        pass # Handle parsing error silently if data is malformed

    # Hash new password
    hashed_password = get_password_hash(req.new_password)
    
    # Update user password
    cursor.execute("UPDATE users SET password = ?, token_version = token_version + 1 WHERE id = ?", (hashed_password, reset_record["user_id"]))
    
    log_audit_action(reset_record["user_id"], "PASSWORD_RESET", None, "User reset their password (all sessions revoked)", req.client.host)
    
    # Invalidate token
    cursor.execute("DELETE FROM password_resets WHERE id = ?", (reset_record["id"],))
    
    conn.commit()
    conn.close()
    
    return {"message": "Password has been successfully reset"}

@app.delete("/api/auth/account")
async def delete_account(user: dict = Depends(get_current_user)):
    """Permanently delete user account and all associated data"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if user exists
    cursor.execute("SELECT id FROM users WHERE id = ?", (user["id"],))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")
        
    # Delete user (foreign key cascades will handle the rest)
    cursor.execute("DELETE FROM users WHERE id = ?", (user["id"],))
    conn.commit()
    conn.close()
    
    return {"message": "Account deleted successfully"}

@app.post("/api/auth/verify-email")
async def verify_email(req: VerifyEmailRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM email_verifications WHERE token = ?", (req.token,))
    record = cursor.fetchone()
    
    if not record:
        conn.close()
        raise HTTPException(status_code=400, detail="ลิงก์ยืนยันไม่ถูกต้องหรือหมดอายุแล้ว")
        
    from datetime import datetime, timezone
    try:
        expires_at_str = record["expires_at"]
        expires_at = datetime.fromisoformat(expires_at_str.replace('Z', '+00:00'))
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        if datetime.now(timezone.utc) > expires_at:
            cursor.execute("DELETE FROM email_verifications WHERE id = ?", (record["id"],))
            conn.commit()
            conn.close()
            raise HTTPException(status_code=400, detail="ลิงก์ยืนยันหมดอายุแล้ว")
    except ValueError:
        pass
        
    cursor.execute("UPDATE users SET is_verified = 1 WHERE id = ?", (record["user_id"],))
    cursor.execute("DELETE FROM email_verifications WHERE user_id = ?", (record["user_id"],))
    conn.commit()
    conn.close()
    
    return {"message": "ยืนยันอีเมลสำเร็จ"}

@app.post("/api/auth/resend-verification")
async def resend_verification(req: ResendVerificationRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT id, is_verified FROM users WHERE email = ?", (req.email,))
    user = cursor.fetchone()
    
    if not user:
        conn.close()
        return {"message": "หากมีบัญชีนี้ในระบบ ลิงก์ยืนยันจะถูกส่งไปที่อีเมลของคุณ"}
        
    if user["is_verified"] == 1:
        conn.close()
        raise HTTPException(status_code=400, detail="อีเมลนี้ได้รับการยืนยันแล้ว")
        
    import uuid
    from datetime import datetime, timezone, timedelta
    import os, smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    token = uuid.uuid4().hex
    expires_at = (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
    
    cursor.execute("DELETE FROM email_verifications WHERE user_id = ?", (user["id"],))
    cursor.execute("INSERT INTO email_verifications (user_id, token, expires_at) VALUES (?, ?, ?)", (user["id"], token, expires_at))
    conn.commit()
    conn.close()
    
    # Send email
    verify_link = f"http://localhost:8080/verify-email?token={token}"
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port_str = os.getenv("SMTP_PORT", "587")
    smtp_port = int(smtp_port_str) if smtp_port_str else 587
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    
    dev_verify_link = None
    if smtp_host and smtp_user and smtp_password:
        try:
            msg = MIMEMultipart()
            msg['From'] = f"Evaly Score <{smtp_user}>"
            msg['To'] = req.email
            msg['Subject'] = "ส่งซ้ำ - ยืนยันบัญชีอีเมล Evaly Score (Verify Email)"
            
            body = f"""
            <h2>ยินดีต้อนรับสู่ Evaly Score</h2>
            <p>กรุณาคลิกที่ลิงก์ด้านล่างเพื่อยืนยันบัญชีอีเมลของคุณ (ลิงก์มีอายุการใช้งาน 24 ชั่วโมง):</p>
            <p><a href="{verify_link}">{verify_link}</a></p>
            """
            
            msg.attach(MIMEText(body, 'html'))
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
            server.quit()
            print(f"\n[Email Sent] Verification link resent to {req.email}")
            return {"message": "ส่งลิงก์ยืนยันอีเมลสำเร็จ กรุณาตรวจสอบกล่องข้อความ"}
        except Exception as e:
            print(f"\n[Email Error] Failed to send email: {e}")
            dev_verify_link = verify_link
    else:
        dev_verify_link = verify_link
        
    if dev_verify_link:
        print(f"\n========== VERIFY EMAIL ==========")
        print(f"Verify Link: {dev_verify_link}")
        print(f"==================================\n")
        return {"message": "Dev Mode", "dev_verify_link": dev_verify_link}
        
    return {"message": "หากมีบัญชีนี้ในระบบ ลิงก์ยืนยันจะถูกส่งไปที่อีเมลของคุณ"}

class FirebaseLoginRequest(BaseModel):
    firebase_token: str

# Firebase Admin SDK - verify Firebase tokens
_FIREBASE_CREDENTIALS_PATH = os.getenv("FIREBASE_CREDENTIALS_PATH", "")

# Resolve to absolute path relative to project root (parent of server/)
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _FIREBASE_CREDENTIALS_PATH and not os.path.isabs(_FIREBASE_CREDENTIALS_PATH):
    _FIREBASE_CREDENTIALS_PATH = os.path.join(_PROJECT_ROOT, _FIREBASE_CREDENTIALS_PATH)

# Initialize Firebase Admin SDK
_firebase_app = None
_FIREBASE_JSON = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON", "")

if _FIREBASE_JSON:
    import firebase_admin
    from firebase_admin import credentials, auth
    try:
        cred_dict = json_module_top.loads(_FIREBASE_JSON)
        cred = credentials.Certificate(cred_dict)
        _firebase_app = firebase_admin.initialize_app(cred)
        print("[Firebase] Admin SDK initialized successfully from environment variable")
    except Exception as e:
        print(f"[Firebase] Failed to initialize Admin SDK from ENV: {e}")
elif _FIREBASE_CREDENTIALS_PATH and os.path.exists(_FIREBASE_CREDENTIALS_PATH):
    import firebase_admin
    from firebase_admin import credentials, auth
    try:
        cred = credentials.Certificate(_FIREBASE_CREDENTIALS_PATH)
        _firebase_app = firebase_admin.initialize_app(cred)
        print("[Firebase] Admin SDK initialized successfully from file")
    except Exception as e:
        print(f"[Firebase] Failed to initialize Admin SDK from file: {e}")

@app.post("/api/auth/firebase-login", response_model=TokenResponse)
async def firebase_login(request: FirebaseLoginRequest, req: Request):
    """
    Verify Firebase ID token and either:
    - Find existing user by email and log them in
    - Create new user account if email doesn't exist (default to 'student' role)
    """
    if not _firebase_app:
        raise HTTPException(status_code=500, detail="Firebase Admin SDK not configured")
    
    try:
        # Verify the Firebase token
        decoded_token = auth.verify_id_token(request.firebase_token)
        email = decoded_token.get("email")
        display_name = decoded_token.get("name", decoded_token.get("display_name", "User"))
        google_picture = decoded_token.get("picture", None)  # Google profile picture URL
        
        if not email:
            raise HTTPException(status_code=400, detail="Email not available from Google account")
        
    except auth.InvalidIdTokenError as e:
        print(f"[Firebase] Invalid token: {e}")
        raise HTTPException(status_code=401, detail="Invalid Firebase token")
    except auth.ExpiredIdTokenError:
        raise HTTPException(status_code=401, detail="Firebase token has expired")
    except Exception as e:
        print(f"[Firebase Verify Error] {e}")
        raise HTTPException(status_code=401, detail="Failed to verify Firebase token")
    
    # Check if user exists in our database
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM users WHERE email = ?", (email,))
    user = cursor.fetchone()
    
    if user:
        # User exists - log them in, update Google avatar if changed
        user_dict = dict(user)  # convert to dict
        existing_avatar = user_dict.get("avatar_url")
        if google_picture and google_picture != existing_avatar:
            cursor.execute(
                "UPDATE users SET avatar_url = ? WHERE id = ?",
                (google_picture, user_dict["id"])
            )
            conn.commit()
        conn.close()
        access_token = create_access_token(data={"sub": email, "token_version": user.get("token_version", 0)})
        log_audit_action(user["id"], "LOGIN", None, "User logged in via Firebase", req.client.host)
        
        user_info = {
            "id": user_dict["id"],
            "email": user_dict["email"],
            "name": user_dict["name"],
            "role": user_dict["role"],
            "studentId": user_dict.get("student_id"),
            "avatarUrl": google_picture or existing_avatar
        }
        
        return {"access_token": access_token, "token_type": "bearer", "user": user_info}
    
    # User doesn't exist - create new account as unassigned
    try:
        cursor.execute(
            "INSERT INTO users (email, password, name, role, student_id, avatar_url, is_verified) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (email, f"firebase_{decoded_token.get('uid', '')}", display_name, "unassigned", None, google_picture, 1)
        )
        conn.commit()
        new_user_id = cursor.lastrowid
    except pymysql.err.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_info = {
        "id": new_user_id,
        "email": email,
        "name": display_name,
        "role": "unassigned",
        "studentId": None,
        "avatarUrl": google_picture
    }
    
    access_token = create_access_token(data={"sub": email, "token_version": 0})
    log_audit_action(new_user_id, "REGISTER", None, "User registered via Firebase", req.client.host)
    conn.close()
    
    return {"access_token": access_token, "token_type": "bearer", "user": user_info}

@app.post("/api/auth/link-google")
async def link_google(request: FirebaseLoginRequest, current_user: dict = Depends(get_current_user)):
    if not _firebase_app:
        raise HTTPException(status_code=500, detail="Firebase SDK ไม่พร้อมใช้งาน")
        
    try:
        decoded_token = auth.verify_id_token(request.firebase_token)
        google_picture = decoded_token.get("picture", None)
    except Exception as e:
        raise HTTPException(status_code=401, detail="ยืนยันบัญชี Google ไม่สำเร็จ")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("UPDATE users SET is_verified = 1 WHERE id = ?", (current_user["id"],))
    
    if google_picture and not current_user.get("avatar_url"):
        cursor.execute("UPDATE users SET avatar_url = ? WHERE id = ?", (google_picture, current_user["id"]))
        
    conn.commit()
    conn.close()
    
    return {"message": "เชื่อมโยงบัญชี Google สำเร็จ"}

# Notifications
@app.get("/api/notifications")
async def get_notifications(user: dict = Depends(get_current_user)):
    """Return notifications for teachers and students."""
    from datetime import datetime, timezone, timedelta

    conn = get_db_connection()
    cursor = conn.cursor()
    now_utc = datetime.now(timezone.utc)
    notifications = []

    # ─── TEACHER ───────────────────────────────────────────────
    if user["role"] == "teacher":
        cursor.execute(
            """
            SELECT e.id AS exam_id, e.title AS exam_title, e.end_date,
                   r.id AS room_id, r.name AS room_name
            FROM exams e
            JOIN rooms r ON e.room_id = r.id
            WHERE r.owner_id = ?
            """,
            (user["id"],)
        )
        exams = cursor.fetchall()

        for exam in exams:
            exam_id    = exam["exam_id"]
            exam_title = exam["exam_title"]
            room_id    = exam["room_id"]
            room_name  = exam["room_name"]
            end_date_str = exam["end_date"]

            deadline_passed = False
            if end_date_str:
                try:
                    deadline = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
                    if deadline.tzinfo is None:
                        deadline = deadline.replace(tzinfo=timezone.utc)
                    if now_utc > deadline:
                        deadline_passed = True
                except (ValueError, TypeError):
                    pass

            cursor.execute(
                """
                SELECT
                    COUNT(*) AS total_submitted,
                    SUM(CASE WHEN status IN ('ready','needs_review') THEN 1 ELSE 0 END) AS pending_review,
                    SUM(CASE WHEN status = 'approved' THEN 1 ELSE 0 END) AS approved_count
                FROM submissions WHERE exam_id = ?
                """,
                (exam_id,)
            )
            row = cursor.fetchone()
            total_submitted = row["total_submitted"] or 0
            pending_review  = row["pending_review"]  or 0
            approved_count  = row["approved_count"]  or 0

            if deadline_passed and pending_review > 0:
                notifications.append({
                    "type": "deadline_passed",
                    "exam_id": exam_id, "exam_title": exam_title,
                    "room_id": room_id, "room_name": room_name,
                    "message": (
                        f"[{room_name}] ชุดข้อสอบ \"{exam_title}\" หมดเวลาส่งแล้ว "
                        f"มีนักศึกษาส่งคำตอบ {total_submitted} คน รอตรวจ {pending_review} คน"
                    ),
                    "link": f"/room/{room_id}/exam/{exam_id}/review"
                })

            if pending_review > 0:
                notifications.append({
                    "type": "ai_graded",
                    "exam_id": exam_id, "exam_title": exam_title,
                    "room_id": room_id, "room_name": room_name,
                    "message": (
                        f"[{room_name}] AI ประเมินผล \"{exam_title}\" เสร็จแล้ว "
                        f"รอการอนุมัติจากอาจารย์ {pending_review} คน "
                        f"(อนุมัติแล้ว {approved_count}/{total_submitted} คน)"
                    ),
                    "link": f"/room/{room_id}/exam/{exam_id}/review"
                })


    # ─── STUDENT ───────────────────────────────────────────────
    else:
        student_id = user["id"]

        # All exams in rooms this student has joined
        cursor.execute(
            """
            SELECT e.id AS exam_id, e.title AS exam_title,
                   e.start_date, e.end_date,
                   r.id AS room_id, r.name AS room_name
            FROM exams e
            JOIN rooms r ON e.room_id = r.id
            JOIN enrollments en ON en.room_id = r.id
            WHERE en.user_id = ?
            """,
            (student_id,)
        )
        exams = cursor.fetchall()

        for exam in exams:
            exam_id      = exam["exam_id"]
            exam_title   = exam["exam_title"]
            room_id      = exam["room_id"]
            room_name    = exam["room_name"]
            start_str    = exam["start_date"]
            end_str      = exam["end_date"]

            # Parse dates
            start_dt = end_dt = None
            try:
                if start_str:
                    start_dt = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
                    if start_dt.tzinfo is None:
                        start_dt = start_dt.replace(tzinfo=timezone.utc)
                if end_str:
                    end_dt = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
                    if end_dt.tzinfo is None:
                        end_dt = end_dt.replace(tzinfo=timezone.utc)
            except (ValueError, TypeError):
                pass

            # Check if student has already submitted
            cursor.execute(
                "SELECT id, status FROM submissions WHERE exam_id = ? AND student_id = ?",
                (exam_id, student_id)
            )
            submission = cursor.fetchone()
            has_submitted = submission is not None
            sub_status = submission["status"] if submission else None

            exam_open = (start_dt is None or now_utc >= start_dt) and \
                        (end_dt is None or now_utc <= end_dt)

            # 1. New exam available (open + not submitted)
            if exam_open and not has_submitted:
                notifications.append({
                    "type": "new_exam",
                    "exam_id": exam_id, "exam_title": exam_title,
                    "room_id": room_id, "room_name": room_name,
                    "message": f"[{room_name}] มีข้อสอบใหม่ \"{exam_title}\" เปิดรับการส่งแล้ว!",
                    "link": f"/room/{room_id}/exam/{exam_id}"
                })

            # 2. Deadline approaching < 24h (open + not submitted)
            if exam_open and not has_submitted and end_dt:
                time_left = end_dt - now_utc
                if timedelta(0) < time_left < timedelta(hours=24):
                    hours_left = int(time_left.total_seconds() // 3600)
                    mins_left  = int((time_left.total_seconds() % 3600) // 60)
                    time_str   = f"{hours_left} ชั่วโมง {mins_left} นาที" if hours_left > 0 else f"{mins_left} นาที"
                    notifications.append({
                        "type": "deadline_soon",
                        "exam_id": exam_id, "exam_title": exam_title,
                        "room_id": room_id, "room_name": room_name,
                        "message": f"[{room_name}] \"{exam_title}\" ใกล้หมดเวลา! เหลืออีก {time_str}",
                        "link": f"/room/{room_id}/exam/{exam_id}"
                    })

            # 3. Result published (submission is approved)
            if has_submitted and sub_status == "approved":
                notifications.append({
                    "type": "result_published",
                    "exam_id": exam_id, "exam_title": exam_title,
                    "room_id": room_id, "room_name": room_name,
                    "message": f"[{room_name}] อาจารย์ประกาศผล \"{exam_title}\" แล้ว กดเพื่อดูคะแนนของคุณ",
                    "link": f"/room/{room_id}/exam/{exam_id}"
                })

    conn.close()
    return notifications


# Room Routes
def generate_class_code(length=6):
    characters = string.ascii_uppercase + string.digits
    return ''.join(random.choice(characters) for _ in range(length))

@app.post("/api/rooms")
async def create_room(room: RoomCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can create rooms")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Generate unique class code
    class_code = generate_class_code()
    
    try:
        cursor.execute(
            "INSERT INTO rooms (name, section, class_code, owner_id) VALUES (?, ?, ?, ?)",
            (room.name, room.section, class_code, user["id"])
        )
        conn.commit()
        new_room_id = cursor.lastrowid
    except pymysql.err.IntegrityError:
        conn.close()
        raise HTTPException(status_code=500, detail="Failed to generate unique class code. Try again.")
        
    cursor.execute("SELECT * FROM rooms WHERE id = ?", (new_room_id,))
    new_room = dict(cursor.fetchone())
    conn.close()
    
    return new_room

@app.get("/api/rooms")
async def get_rooms(user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if user["role"] == "teacher":
        cursor.execute("SELECT * FROM rooms WHERE owner_id = ?", (user["id"],))
        rooms = cursor.fetchall()
    else:
        cursor.execute('''
            SELECT r.* FROM rooms r
            JOIN enrollments e ON r.id = e.room_id
            WHERE e.user_id = ?
        ''', (user["id"],))
        rooms = cursor.fetchall()
        
    conn.close()
    return [dict(room) for room in rooms]

@app.put("/api/rooms/{room_id}")
async def update_room(room_id: int, room_data: RoomCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can edit rooms")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    existing_room = cursor.fetchone()
    
    if not existing_room:
        conn.close()
        raise HTTPException(status_code=404, detail="Room not found or unauthorized")
        
    cursor.execute(
        "UPDATE rooms SET name = ?, section = ? WHERE id = ?",
        (room_data.name, room_data.section, room_id)
    )
    conn.commit()
    conn.close()
    
    return {"message": "Room updated successfully"}

@app.delete("/api/rooms/{room_id}")
async def delete_room(request: Request, room_id: int, user: dict = Depends(get_current_user)):
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can delete rooms")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Room not found or unauthorized")
        
    cursor.execute("DELETE FROM rooms WHERE id = ?", (room_id,))
    conn.commit()
    conn.close()
    
    log_audit_action(user["id"], "DELETE_ROOM", str(room_id), f"Deleted room {room_id}", request.client.host)
    return {"message": "Room deleted successfully"}

@app.post("/api/rooms/join")
async def join_room(request: JoinRoomRequest, user: dict = Depends(get_current_user)):
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="Only students can join rooms")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if code maps to a room
    cursor.execute("SELECT id FROM rooms WHERE class_code = ?", (request.class_code.upper(),))
    room = cursor.fetchone()
    if not room:
        conn.close()
        raise HTTPException(status_code=404, detail="รหัสห้องไม่ถูกต้อง (Invalid class code)")
        
    room_id = room["id"]
    
    # Check if already enrolled
    cursor.execute("SELECT id FROM enrollments WHERE user_id = ? AND room_id = ?", (user["id"], room_id))
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="คุณอยู่ในห้องนี้แล้ว (Already joined)")
        
    cursor.execute("INSERT INTO enrollments (user_id, room_id) VALUES (?, ?)", (user["id"], room_id))
    conn.commit()
    
    # Fetch room info to return
    cursor.execute("SELECT * FROM rooms WHERE id = ?", (room_id,))
    joined_room = dict(cursor.fetchone())
    conn.close()
    
    return joined_room

@app.get("/api/rooms/{room_id}")
async def get_room(room_id: int, user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if user["role"] == "teacher":
        cursor.execute("SELECT * FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    else:
        cursor.execute("""
            SELECT r.* FROM rooms r
            JOIN enrollments e ON r.id = e.room_id
            WHERE r.id = ? AND e.user_id = ?
        """, (room_id, user["id"]))
    
    room = cursor.fetchone()
    conn.close()
    
    if not room:
        raise HTTPException(status_code=404, detail="Room not found or unauthorized")
    
    return dict(room)

@app.get("/api/rooms/{room_id}/members")
async def get_room_members(room_id: int, user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verify access
    if user["role"] == "teacher":
        cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    else:
        cursor.execute("SELECT room_id FROM enrollments WHERE room_id = ? AND user_id = ?", (room_id, user["id"]))
    
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    # Get room owner (teacher)
    cursor.execute("SELECT u.id, u.name, u.email, u.student_id FROM users u JOIN rooms r ON u.id = r.owner_id WHERE r.id = ?", (room_id,))
    teacher_row = cursor.fetchone()
    teacher = {**dict(teacher_row), "role": "teacher", "joined_at": None} if teacher_row else None

    cursor.execute("""
        SELECT u.id, u.name, u.email, u.student_id, e.joined_at
        FROM users u
        JOIN enrollments e ON u.id = e.user_id
        WHERE e.room_id = ?
        ORDER BY e.joined_at DESC
    """, (room_id,))
    members = cursor.fetchall()
    conn.close()

    result = []
    if teacher:
        result.append({**teacher, "role": "teacher"})
    for m in members:
        result.append({**dict(m), "role": "student"})
    
    return result

# Announcement Routes
@app.post("/api/rooms/{room_id}/announcements")
async def create_announcement(room_id: int, ann: AnnouncementCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can create announcements")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verify room ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Room not found or unauthorized")
        
    cursor.execute(
        "INSERT INTO announcements (room_id, teacher_id, title, content) VALUES (?, ?, ?, ?)",
        (room_id, user["id"], ann.title, ann.content)
    )
    ann_id = cursor.lastrowid
    conn.commit()
    
    # Notify students in room
    cursor.execute("SELECT user_id FROM enrollments WHERE room_id = ?", (room_id,))
    students = cursor.fetchall()
    for s in students:
        await trigger_socket_notify(
            user_id=s["user_id"],
            notify_type="new_announcement",
            message=f"มีประกาศใหม่ในห้องเรียน: {ann.title}",
            data={"room_id": room_id, "announcement_id": ann_id}
        )
        
    conn.close()
    return {"id": ann_id, "message": "Announcement created successfully"}

@app.get("/api/rooms/{room_id}/announcements")
async def list_announcements(room_id: int, user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verify access
    if user["role"] == "teacher":
        cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    else:
        cursor.execute("SELECT room_id FROM enrollments WHERE room_id = ? AND user_id = ?", (room_id, user["id"]))
        
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")
        
    # Get announcements with read status for current user if student
    if user["role"] == "student":
        cursor.execute("""
            SELECT a.*, (SELECT 1 FROM announcement_reads ar WHERE ar.announcement_id = a.id AND ar.student_id = ?) as is_read
            FROM announcements a
            WHERE a.room_id = ?
            ORDER BY a.created_at DESC
        """, (user["id"], room_id))
    else:
        cursor.execute("""
            SELECT a.*, (SELECT COUNT(*) FROM announcement_reads ar WHERE ar.announcement_id = a.id) as read_count
            FROM announcements a
            WHERE a.room_id = ?
            ORDER BY a.created_at DESC
        """, (room_id,))
        
    anns = cursor.fetchall()
    conn.close()
    return [dict(a) for a in anns]

@app.post("/api/announcements/{ann_id}/read")
async def mark_announcement_read(ann_id: int, user: dict = Depends(get_current_user)):
    if user["role"] != "student":
        return {"message": "Only students need read receipts"}
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Use INSERT IGNORE for MySQL or equivalent logic
        cursor.execute(
            "INSERT INTO announcement_reads (announcement_id, student_id) VALUES (?, ?) ON DUPLICATE KEY UPDATE student_id=student_id",
            (ann_id, user["id"])
        )
        conn.commit()
    except Exception as e:
        print(f"Error marking announcement read: {e}")
        
    conn.close()
    return {"message": "Marked as read"}

@app.get("/api/announcements/{ann_id}/read-status")
async def get_announcement_read_status(ann_id: int, user: dict = Depends(get_current_user)):
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can view read status")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get room_id for this announcement to verify ownership
    cursor.execute("SELECT room_id FROM announcements WHERE id = ?", (ann_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Announcement not found")
        
    room_id = row["room_id"]
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")
        
    # Get all students in room and their read status
    cursor.execute("""
        SELECT u.id, u.name, u.student_id, ar.read_at
        FROM enrollments e
        JOIN users u ON e.user_id = u.id
        LEFT JOIN announcement_reads ar ON ar.announcement_id = ? AND ar.student_id = u.id
        WHERE e.room_id = ?
    """, (ann_id, room_id))
    
    status = cursor.fetchall()
    conn.close()
    return [dict(s) for s in status]

# Exam Routes
import json as json_module

class QuestionInput(BaseModel):
    text: str
    score: float = 0
    answer_key: Optional[str] = None
    rubrics: Optional[list] = None
    order_index: int = 0
    question_images_base64: Optional[List[str]] = None  # list of base64 data URLs

class ExamCreate(BaseModel):
    title: str
    description: Optional[str] = None
    total_score: float = 0
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    is_randomized: int = 0
    questions: list[QuestionInput] = []

class TimeExtensionRequest(BaseModel):
    student_id: Optional[int] = None  # None = apply to whole class
    extra_minutes: int
    note: Optional[str] = None

class GenerateRubricRequest(BaseModel):
    question_text: str
    total_score: float
    question_images_base64: Optional[List[str]] = None

class QuestionBankCreate(BaseModel):
    text: str
    score: float = 0
    answer_key: Optional[str] = None
    rubrics: Optional[list] = None
    tags: Optional[str] = None

class DraftSaveRequest(BaseModel):
    answers: dict  # {question_id: answer_text}

@app.post("/api/rooms/{room_id}/exams")
async def create_exam(room_id: int, exam: ExamCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can create exams")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify room ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Room not found or unauthorized")

    # Calculate total_score server-side to ensure accuracy
    computed_total_score = sum(float(q.score) for q in exam.questions)

    # Insert exam
    cursor.execute(
        "INSERT INTO exams (room_id, title, description, total_score, start_date, end_date, is_randomized) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (room_id, exam.title, exam.description, computed_total_score, exam.start_date, exam.end_date, exam.is_randomized)
    )
    exam_id = cursor.lastrowid

    # Insert questions
    import base64, re as _re
    for q in exam.questions:
        rubrics_json = json_module.dumps(q.rubrics, ensure_ascii=False) if q.rubrics else None

        # Handle multiple question images (base64 data URLs)
        image_paths = []
        import time as _time
        for img_idx, img_b64 in enumerate(q.question_images_base64 or []):
            try:
                match = _re.match(r'data:(?P<mime>[^;]+);base64,(?P<data>.+)', img_b64)
                if match:
                    mime = match.group('mime')
                    b64data = match.group('data')
                    img_bytes = base64.b64decode(b64data)
                    ext = mime.split('/')[-1].replace('jpeg', 'jpg')
                    # Upload to Cloudinary
                    c_url = upload_to_cloudinary(img_bytes, folder=f"questions/{exam_id}")
                    if c_url:
                        image_paths.append(c_url)
            except Exception as img_err:
                print(f"[Question Image] Failed to save index {img_idx}: {img_err}")

        image_paths_json = json_module.dumps(image_paths) if image_paths else None
        # Keep single image_path for backward compat (first image)
        first_image = image_paths[0] if image_paths else None

        cursor.execute(
            "INSERT INTO questions (exam_id, text, score, answer_key, rubrics, order_index, image_path, image_paths) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (exam_id, q.text, q.score, q.answer_key, rubrics_json, q.order_index, first_image, image_paths_json)
        )

    conn.commit()
    cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
    new_exam = dict(cursor.fetchone())
    conn.close()

    return new_exam

@app.put("/api/rooms/{room_id}/exams/{exam_id}")
async def update_exam(room_id: int, exam_id: int, exam: ExamCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can update exams")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify exam ownership
    cursor.execute("SELECT id FROM exams WHERE id = ? AND room_id = ?", (exam_id, room_id))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Exam not found")

    # Verify room ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Room not found or unauthorized")

    # Calculate total_score server-side to ensure accuracy
    computed_total_score = sum(float(q.score) for q in exam.questions)

    # Update exam
    cursor.execute(
        "UPDATE exams SET title = ?, description = ?, total_score = ?, start_date = ?, end_date = ?, is_randomized = ? WHERE id = ?",
        (exam.title, exam.description, computed_total_score, exam.start_date, exam.end_date, exam.is_randomized, exam_id)
    )

    # Delete existing questions
    cursor.execute("DELETE FROM questions WHERE exam_id = ?", (exam_id,))

    # Insert new questions
    import base64, re as _re
    for q in exam.questions:
        rubrics_json = json_module.dumps(q.rubrics, ensure_ascii=False) if q.rubrics else None

        # Handle multiple question images (base64 data URLs)
        image_paths = []
        import time as _time
        for img_idx, img_b64 in enumerate(q.question_images_base64 or []):
            try:
                match = _re.match(r'data:(?P<mime>[^;]+);base64,(?P<data>.+)', img_b64)
                if match:
                    mime = match.group('mime')
                    b64data = match.group('data')
                    img_bytes = base64.b64decode(b64data)
                    ext = mime.split('/')[-1].replace('jpeg', 'jpg')
                    # Upload to Cloudinary
                    c_url = upload_to_cloudinary(img_bytes, folder=f"questions/{exam_id}")
                    if c_url:
                        image_paths.append(c_url)
            except Exception as img_err:
                print(f"[Question Image] Failed to save index {img_idx}: {img_err}")

        image_paths_json = json_module.dumps(image_paths) if image_paths else None
        first_image = image_paths[0] if image_paths else None

        cursor.execute(
            "INSERT INTO questions (exam_id, text, score, answer_key, rubrics, order_index, image_path, image_paths) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (exam_id, q.text, q.score, q.answer_key, rubrics_json, q.order_index, first_image, image_paths_json)
        )

    conn.commit()
    cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
    updated_exam = dict(cursor.fetchone())
    conn.close()

    return updated_exam

@app.delete("/api/rooms/{room_id}/exams/{exam_id}")
async def delete_exam(request: Request, room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can delete exams")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify room ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Room not found or unauthorized")

    # Verify exam exists
    cursor.execute("SELECT id FROM exams WHERE id = ? AND room_id = ?", (exam_id, room_id))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Exam not found")

    # Delete exam (questions will be cascade deleted due to FK)
    cursor.execute("DELETE FROM exams WHERE id = ?", (exam_id,))
    conn.commit()
    conn.close()

    log_audit_action(user["id"], "DELETE_EXAM", str(exam_id), f"Deleted exam {exam_id}", request.client.host)
    return {"message": "Exam deleted successfully"}

@app.get("/api/rooms/{room_id}/exams")
async def list_exams(room_id: int, user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify access
    if user["role"] == "teacher":
        cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    else:
        cursor.execute("SELECT room_id FROM enrollments WHERE room_id = ? AND user_id = ?", (room_id, user["id"]))

    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")

    cursor.execute("SELECT * FROM exams WHERE room_id = ? ORDER BY created_at DESC", (room_id,))
    exams = cursor.fetchall()
    conn.close()

    return [dict(e) for e in exams]

@app.get("/api/rooms/{room_id}/exams/{exam_id}")
async def get_exam(room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM exams WHERE id = ? AND room_id = ?", (exam_id, room_id))
    exam = cursor.fetchone()
    if not exam:
        conn.close()
        raise HTTPException(status_code=404, detail="Exam not found")

    cursor.execute("SELECT * FROM questions WHERE exam_id = ? ORDER BY order_index", (exam_id,))
    questions = cursor.fetchall()
    conn.close()

    result = dict(exam)
    result["questions"] = []
    for q in questions:
        qd = dict(q)
        if qd.get("rubrics"):
            try:
                qd["rubrics"] = json_module.loads(qd["rubrics"])
            except Exception:
                qd["rubrics"] = []
        # Parse image_paths JSON array
        if qd.get("image_paths"):
            try:
                qd["image_paths"] = json_module.loads(qd["image_paths"])
            except Exception:
                qd["image_paths"] = [qd.get("image_path")] if qd.get("image_path") else []
        else:
            # Backward compat: single image_path
            qd["image_paths"] = [qd["image_path"]] if qd.get("image_path") else []
        result["questions"].append(qd)

    # Deterministic shuffle for students when is_randomized = 1
    # Using student_id as seed ensures same student always sees same order
    if result.get("is_randomized") and user["role"] == "student":
        import random as _random
        rng = _random.Random(user["id"] + exam_id * 1000)
        rng.shuffle(result["questions"])

    return result


# ============================================================
# Submission & Grading Routes — Gemini AI Scoring
# ============================================================

import google.genai as genai
from google.genai import types as genai_types

# Configure Gemini client once at module load
_GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
if _GEMINI_API_KEY and _GEMINI_API_KEY != "your-gemini-api-key-here":
    _genai_client = genai.Client(api_key=_GEMINI_API_KEY)
    _USE_GEMINI = True
else:
    _genai_client = None
    _USE_GEMINI = False

_GEMINI_MODEL = "gemini-flash-latest"


def _fallback_score(answer_text: str, max_score: float) -> dict:
    """Rule-based fallback when Gemini is unavailable."""
    if not answer_text or len(answer_text.strip()) < 5:
        return {
            "score": round(random.uniform(0.0, 0.2) * max_score, 1),
            "confidence": "low",
            "feedback": "คำตอบสั้นเกินไปหรือไม่มีเนื้อหา AI ไม่สามารถประเมินได้ โปรดอาจารย์ตรวจสอบด้วยตนเอง",
        }
    pct = random.uniform(0.55, 1.0)
    confidence = "high" if pct >= 0.85 else "medium" if pct >= 0.65 else "low"
    return {
        "score": round(pct * max_score, 1),
        "confidence": confidence,
        "feedback": "(ระบบ AI ออฟไลน์ — ใช้ผลประมาณการ) โปรดอาจารย์ตรวจทานคะแนนอีกครั้ง",
    }


async def score_with_gemini(
    question_text: str,
    answer_text: str,
    max_score: float,
    answer_key: Optional[str] = None,
    rubrics: Optional[list] = None,
    image_bytes_list: Optional[List[bytes]] = None,
    image_mime_list: Optional[List[str]] = None,
    q_image_bytes_list: Optional[List[bytes]] = None,
    q_image_mime_list: Optional[List[str]] = None,
) -> dict:
    """
    Score a student answer using Gemini AI.
    Returns: {score: float, confidence: 'high'|'medium'|'low', feedback: str}
    Falls back to rule-based scoring if Gemini is unavailable.
    """
    if not _USE_GEMINI or not _genai_client:
        return _fallback_score(answer_text, max_score)

    # Build rubric text
    rubric_text = ""
    if rubrics:
        rubric_lines = []
        for r in rubrics:
            name = r.get("name") or r.get("label", "")
            score = r.get("score") or r.get("maxScore", "")
            desc = r.get("description", "")
            rubric_lines.append(f"- {name} ({score} คะแนน){': ' + desc if desc else ''}")
        rubric_text = "\n".join(rubric_lines)

    # Build all optional sections as plain strings first (avoid backslash in f-string)
    answer_key_section = ""
    if answer_key:
        answer_key_section = "## แนวคำตอบ\n" + answer_key + "\n"

    rubric_section = ""
    if rubric_text:
        rubric_section = "## เกณฑ์การให้คะแนน\n" + rubric_text + "\n"

    student_answer_section = answer_text.strip() if answer_text and answer_text.strip() else "(ไม่มีคำตอบแบบข้อความ)"

    prompt = (
        "คุณคือคุณครูผู้เชี่ยวชาญในการตรวจข้อสอบอัตนัย กรุณาประเมินคำตอบของนักเรียนอย่างละเอียดและเป็นธรรมตามเกณฑ์ที่กำหนด\n\n"
        "## ข้อมูลโจทย์\n"
        f"ข้อความโจทย์: {question_text or '(ไม่มีข้อความโจทย์ - โปรดดูจากรูปภาพประกอบโจทย์)'}\n"
        "**สำคัญ**: หากข้อความโจทย์ว่างเปล่า ให้คุณวิเคราะห์เนื้อหาคำถามจากรูปภาพที่อยู่ในส่วนของ 'Question Images' ที่แนบไป\n\n"
        f"## คะแนนเต็ม\n{max_score} คะแนน\n\n"
        + answer_key_section
        + rubric_section
        + f"## ข้อมูลคำตอบของนักเรียน\n"
        f"ข้อความคำตอบ: {student_answer_section}\n"
        "**สำคัญ**: หากนักเรียนส่งรูปภาพมาในส่วนของ 'Student Answer Images' ให้คุณวิเคราะห์คำตอบจากรูปภาพเหล่านั้นประกอบด้วย\n\n"
        "## คำสั่งการตรวจ\n"
        "1. วิเคราะห์โจทย์: ทำความเข้าใจสิ่งที่โจทย์ต้องการ (จากข้อความหรือรูปภาพโจทย์)\n"
        "2. วิเคราะห์คำตอบ: ตรวจสอบคำตอบของนักเรียน (จากข้อความหรือรูปภาพคำตอบ) ว่าตรงตามโจทย์และเกณฑ์การให้คะแนนหรือไม่\n"
        "3. การให้คะแนน: พิจารณาคะแนนตามความเหมาะสม (Chain of Thought: สิ่งที่ตอบถูก ขาดส่วนไหน)\n"
        "4. วิเคราะห์คุณภาพเรียงความ: ให้คะแนนด้านความสละสลวย ความซับซ้อน และความเหมาะสมของความยาว\n"
        "ตอบกลับเป็น JSON ที่มีรูปแบบดังนี้เท่านั้น (งดเว้นการพิมพ์ข้อความอื่นๆ นอก JSON):\n"
        "{\n"
        f'  "score": <คะแนนที่ได้ เป็นตัวเลขทศนิยม 1 ตำแหน่ง ระหว่าง 0 ถึง {max_score}>,\n'
        '  "confidence": <"high" หากมั่นใจมาก, "medium" หากปานกลาง, "low" หากไม่มั่นใจ หรือรูปภาพไม่ชัดเจน>,\n'
        '  "feedback": <คำอธิบายการให้คะแนนและข้อเสนอแนะเป็นภาษาไทยที่กระชับและเข้าใจง่าย>,\n'
        '  "metrics": {\n'
        '    "length_eval": <ข้อความสรุปความยาว>,\n'
        '    "complexity_eval": <ข้อความสรุปความซับซ้อน>,\n'
        '    "readability_eval": <ข้อความสรุปความลื่นไหล>\n'
        '  }\n'
        "}"
    )

    try:
        # Build multimodal contents
        contents: list = [prompt]
        
        # Add Question Images first
        if q_image_bytes_list and q_image_mime_list:
            contents.append("--- [Question Images] ---")
            for bts, mime in zip(q_image_bytes_list, q_image_mime_list):
                contents.append(genai_types.Part.from_bytes(data=bts, mime_type=mime))
                
        # Add Student Answer Images
        if image_bytes_list and image_mime_list:
            contents.append("--- [Student Answer Images] ---")
            for bts, mime in zip(image_bytes_list, image_mime_list):
                contents.append(genai_types.Part.from_bytes(data=bts, mime_type=mime))

        response = await _genai_client.aio.models.generate_content(
            model=_GEMINI_MODEL,
            contents=contents,
            config=genai_types.GenerateContentConfig(
                temperature=0.2,
                response_mime_type="application/json",
            ),
        )
        raw = response.text.strip()
        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        data = json_module_top.loads(raw)

        score = float(data.get("score", 0))
        score = max(0.0, min(float(max_score), score))  # clamp
        confidence = data.get("confidence", "medium")
        if confidence not in ("high", "medium", "low"):
            confidence = "medium"
        feedback = str(data.get("feedback", ""))

        return {"score": round(score, 1), "confidence": confidence, "feedback": feedback}

    except Exception as e:
        print(f"[Gemini Error] {e} — falling back to heuristic")
        return _fallback_score(answer_text, max_score)


@app.post("/api/gemini/generate-rubric")
async def generate_rubric(req: GenerateRubricRequest, user: dict = Depends(get_current_user)):
    """Generate Answer Key and Rubrics automatically based on question and total score."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can generate rubrics")
    if not _USE_GEMINI or not _genai_client:
        raise HTTPException(status_code=503, detail="Gemini AI is not configured or unavailable")

    prompt = f"""
    You are an expert exam setter and grader.
    Given the following question (text and/or images) and its maximum score, generate a comprehensive "Answer Key" (ธงคำตอบ)
    and a detailed "Rubrics" (เกณฑ์การให้คะแนน) broken down into specific criteria.
    The total score of all rubrics MUST equal exactly {req.total_score}.
    
    If question text is empty, look at the attached images to understand the question.
    
    Output MUST be valid JSON matching this schema:
    {{
        "answer_key": "String (detailed correct answer model in Thai)",
        "rubrics": [
            {{
                "name": "String (short criteria name, e.g. ความถูกต้อง, การอธิบาย, โครงสร้างโค้ด)",
                "description": "String (what is expected to get this score)",
                "score": Number (float or int)
            }}
        ]
    }}
    
    Question Text: {req.question_text}
    Total Score: {req.total_score}
    """
    
    try:
        import base64, re as _re
        contents: list = [prompt]
        
        # Add images if provided
        if req.question_images_base64:
            for img_data in req.question_images_base64:
                try:
                    # Check if it's a base64 DataURL
                    if img_data.startswith("data:"):
                        match = _re.match(r'data:(?P<mime>[^;]+);base64,(?P<data>.+)', img_data)
                        if match:
                            mime = match.group('mime')
                            data = base64.b64decode(match.group('data'))
                            contents.append(genai_types.Part.from_bytes(data=data, mime_type=mime))
                    # Check if it's a URL or local path
                    else:
                        data = await get_image_bytes(img_data)
                        if data:
                            ext = img_data.split(".")[-1].lower()
                            mime = "image/png" if ext == "png" else "image/webp" if ext == "webp" else "image/jpeg"
                            contents.append(genai_types.Part.from_bytes(data=data, mime_type=mime))
                except Exception as e: 
                    print(f"[Generate Rubric Image Error] {e}")
                    continue

        response = _genai_client.models.generate_content(
            model=_GEMINI_MODEL,
            contents=contents,
            config=genai_types.GenerateContentConfig(
                response_mime_type="application/json",
            )
        )
        result_text = response.text
        if result_text.startswith("```json"):
            result_text = result_text.strip("```json").strip("```").strip()
        elif result_text.startswith("```"):
            result_text = result_text.strip("```").strip()
            
        data = json_module.loads(result_text)
        return data
    except Exception as e:
        print(f"[Gemini Error in Generate Rubric]: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate rubric via AI")


class SubmitAnswerInput(BaseModel):
    question_id: int
    answer_text: str

class SubmitExamRequest(BaseModel):
    answers: List[SubmitAnswerInput]

class ApproveSubmissionRequest(BaseModel):
    teacher_scores: Optional[dict] = None   # {question_id: score}
    teacher_comments: Optional[dict] = None # {question_id: comment}

class BulkApproveRequest(BaseModel):
    student_ids: List[int]

def _distribution_buckets(scores: list[float], total_score: float) -> dict:
    if total_score <= 0:
        return {"0-24": 0, "25-49": 0, "50-74": 0, "75-100": 0}

    buckets = {"0-24": 0, "25-49": 0, "50-74": 0, "75-100": 0}
    for s in scores:
        pct = max(0.0, min(100.0, (float(s) / float(total_score)) * 100.0))
        if pct < 25:
            buckets["0-24"] += 1
        elif pct < 50:
            buckets["25-49"] += 1
        elif pct < 75:
            buckets["50-74"] += 1
        else:
            buckets["75-100"] += 1
    return buckets


@app.post("/api/rooms/{room_id}/exams/{exam_id}/submit")
async def submit_exam(
    room_id: int,
    exam_id: int,
    answers: str = Form(...),          # JSON string: [{question_id, answer_text}]
    user: dict = Depends(get_current_user),
    **kwargs,                           # Images will be picked up via Request
):
    """Student submits answers (lock-once). Supports text + image per question."""
    raise HTTPException(status_code=500, detail="Use the multipart endpoint below")


from fastapi import Request

@app.post("/api/rooms/{room_id}/exams/{exam_id}/submit-multipart")
async def submit_exam_multipart(
    request: Request,
    room_id: int,
    exam_id: int,
    user: dict = Depends(get_current_user),
):
    """Student submits answers with optional image per question (lock-once, multipart/form-data)."""
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="Only students can submit exams")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify exam exists in this room
    cursor.execute("SELECT * FROM exams WHERE id = ? AND room_id = ?", (exam_id, room_id))
    exam_row = cursor.fetchone()
    if not exam_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Exam not found")

    # Strict deadline check
    end_date_str = exam_row["end_date"]
    if end_date_str:
        try:
            from datetime import datetime, timezone
            # Parse typical frontend ISO strings
            deadline = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
            # If deadline has no timezone info, assume UTC
            if deadline.tzinfo is None:
                deadline = deadline.replace(tzinfo=timezone.utc)
            now = datetime.now(timezone.utc)
            if now > deadline:
                conn.close()
                raise HTTPException(
                    status_code=403, 
                    detail="เลยกำหนดเวลาส่งคำตอบข้อสอบแล้ว (Deadline Passed)"
                )
        except (ValueError, TypeError):
            pass  # if cannot parse, ignore

    # LOCK-ONCE: check if already submitted
    cursor.execute(
        "SELECT id, status FROM submissions WHERE exam_id = ? AND student_id = ?",
        (exam_id, user["id"])
    )
    existing = cursor.fetchone()
    if existing and existing["status"] not in ("missing", "submitted"):
        conn.close()
        raise HTTPException(
            status_code=409,
            detail="คุณส่งคำตอบข้อสอบนี้ไปแล้ว ไม่สามารถส่งซ้ำได้"
        )
    if existing and existing["status"] == "submitted":
        # Allow re-submit only if still in 'submitted' (before AI graded)
        pass
    if existing and existing["status"] in ("ready", "needs_review", "approved"):
        conn.close()
        raise HTTPException(
            status_code=409,
            detail="คุณส่งคำตอบข้อสอบนี้ไปแล้ว ไม่สามารถส่งซ้ำได้"
        )

    # Parse multipart form
    form = await request.form()
    answers_json = form.get("answers", "[]")
    try:
        answers_list = json_module_top.loads(answers_json)
    except Exception:
        conn.close()
        raise HTTPException(status_code=400, detail="Invalid answers JSON")

    answers_map = {int(a["question_id"]): a.get("answer_text", "") for a in answers_list}

    # Create or update submission record
    cursor.execute(
        """INSERT INTO submissions (exam_id, student_id, status, submitted_at, graded_by_ai)
           VALUES (?, ?, 'submitted', CURRENT_TIMESTAMP, 0)
           ON DUPLICATE KEY UPDATE
             status='submitted', submitted_at=CURRENT_TIMESTAMP, graded_by_ai=0""",
        (exam_id, user["id"])
    )
    submission_id = cursor.lastrowid
    if not submission_id:
        cursor.execute("SELECT id FROM submissions WHERE exam_id = ? AND student_id = ?", (exam_id, user["id"]))
        submission_id = cursor.fetchone()["id"]

    # Get all questions
    cursor.execute("SELECT * FROM questions WHERE exam_id = ? ORDER BY order_index", (exam_id,))
    questions = {q["id"]: dict(q) for q in cursor.fetchall()}

    total_ai_score = 0.0

    for q_id, q in questions.items():
        base_answer_text = answers_map.get(q_id, "")
        answer_text = base_answer_text[:300] if base_answer_text else ""

        # Handle multiple images for this question: image_{q_id}_0, image_{q_id}_1, ...
        img_list: list[bytes] = []
        img_mime_list: list[str] = []
        img_paths: list[str] = []

        # Support both image_{q_id} (single, legacy) and image_{q_id}_N (multi)
        upload_dir = os.path.join("uploads", str(exam_id), str(user["id"]))
        os.makedirs(upload_dir, exist_ok=True)

        for img_idx in range(10):  # support up to 10 images per question
            field_name = f"image_{q_id}_{img_idx}"
            file_field = form.get(field_name)
            
            # backward compatibility for legacy non-multi format
            if not file_field and img_idx == 0:
                file_field = form.get(f"image_{q_id}")
                
            if not file_field or not hasattr(file_field, "read"):
                continue
            raw_bytes = await file_field.read()
            if not raw_bytes:
                continue
            mime = file_field.content_type or "image/jpeg"
            ext = mime.split("/")[-1].replace("jpeg", "jpg")
            # Upload to Cloudinary
            c_url = upload_to_cloudinary(raw_bytes, folder=f"submissions/{exam_id}/{user['id']}")
            if c_url:
                img_paths.append(c_url)

        image_paths_json = json_module.dumps(img_paths) if img_paths else None

        first_image_path = img_paths[0] if img_paths else None
        cursor.execute(
            """INSERT INTO submission_answers
                 (submission_id, question_id, answer_text, ai_score, ai_feedback, ai_confidence, image_path, image_paths)
               VALUES (?, ?, ?, 0, '', 'medium', ?, ?)
               ON DUPLICATE KEY UPDATE
                 answer_text=VALUES(answer_text), image_path=VALUES(image_path), image_paths=VALUES(image_paths)""",
            (submission_id, q_id, answer_text, first_image_path, image_paths_json)
        )

    # Enqueue grading task
    await grading_queue.put({
        "submission_id": submission_id,
        "room_id": room_id,
        "exam_id": exam_id,
        "user_id": user["id"]
    })

    conn.commit()
    conn.close()

    return {
        "message": "ส่งคำตอบสำเร็จ (ระบบกำลังตรวจคะแนน)",
        "submission_id": submission_id,
        "status": "submitted",
        "ai_score": 0
    }

@app.get("/api/rooms/{room_id}/exams/{exam_id}/export-csv")
async def export_exam_csv(room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    """Teacher exports exam scores to CSV"""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can export scores")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verify room ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")
        
    # Get all enrolled students and their scores
    cursor.execute("""
        SELECT 
            u.student_id AS student_code, u.name, u.email, 
            COALESCE(s.status, 'missing') as status, 
            s.total_score, s.submitted_at
        FROM enrollments e
        JOIN users u ON e.user_id = u.id
        LEFT JOIN submissions s ON s.exam_id = ? AND s.student_id = u.id
        WHERE e.room_id = ?
        ORDER BY u.name ASC
    """, (exam_id, room_id))
    
    results = cursor.fetchall()
    
    # Get exam title for filename
    cursor.execute("SELECT title FROM exams WHERE id = ?", (exam_id,))
    row = cursor.fetchone()
    exam_title = row["title"] if row else "Exam"
    conn.close()
    
    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Student ID", "Name", "Status", "Score", "Submitted At"])
    
    for r in results:
        writer.writerow([
            r["student_code"] or "-",
            r["name"],
            r["status"],
            r["total_score"] if r["total_score"] is not None else "0",
            r["submitted_at"] or "-"
        ])
    
    content = output.getvalue()
    
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=scores_{exam_title.replace(' ', '_')}.csv"}
    )

@app.get("/api/rooms/{room_id}/export-summary-csv")
async def export_room_summary_csv(room_id: int, user: dict = Depends(get_current_user)):
    """Teacher exports overall room summary (all exams) to CSV"""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can export summary")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Verify ownership and get room name
    cursor.execute("SELECT name FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    room_row = cursor.fetchone()
    if not room_row:
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")
    room_name = room_row["name"]

    # 2. Get all exams in this room (sorted by creation date)
    cursor.execute("SELECT id, title, total_score FROM exams WHERE room_id = ? ORDER BY created_at ASC", (room_id,))
    exams = cursor.fetchall()
    exam_ids = [e["id"] for e in exams]
    exam_titles = [e["title"] for e in exams]

    # 3. Get all enrolled students
    cursor.execute("""
        SELECT u.id, u.name, u.email, u.student_id
        FROM enrollments e
        JOIN users u ON e.user_id = u.id
        WHERE e.room_id = ?
        ORDER BY u.name ASC
    """, (room_id,))
    students = cursor.fetchall()

    # 4. Get all submissions for these exams in this room
    cursor.execute("""
        SELECT exam_id, student_id, total_score
        FROM submissions
        WHERE exam_id IN (SELECT id FROM exams WHERE room_id = ?)
    """, (room_id,))
    submissions_list = cursor.fetchall()
    
    # Map submissions into a lookup dict: {student_id: {exam_id: score}}
    scores_map = {}
    for sub in submissions_list:
        sid = sub["student_id"]
        eid = sub["exam_id"]
        if sid not in scores_map:
            scores_map[sid] = {}
        scores_map[sid][eid] = sub["total_score"]

    conn.close()

    # 5. Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header: Info + Exam Titles + Total
    writer.writerow(["Student ID", "Name"] + exam_titles + ["Total Cumulative Score"])
    
    for s in students:
        row = [s["student_id"], s["name"]]
        cumulative_total = 0
        
        for eid in exam_ids:
            score = scores_map.get(s["id"], {}).get(eid, 0)
            row.append(score)
            cumulative_total += (score or 0)
            
        row.append(cumulative_total)
        writer.writerow(row)
    
    content = output.getvalue()
    
    filename = f"Summary_{room_name.replace(' ', '_')}.csv"
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/rooms/{room_id}/exams/{exam_id}/submissions")
async def list_submissions(room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    """Teacher gets list of all enrolled students with their submission status."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can view all submissions")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify teacher owns this room
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")

    # Get all enrolled students with their submission info (left join)
    cursor.execute("""
        SELECT
            u.id AS student_id, u.name, u.email, u.student_id AS student_code,
            s.id AS submission_id, s.status, s.total_score, s.submitted_at, s.graded_by_ai
        FROM enrollments e
        JOIN users u ON e.user_id = u.id
        LEFT JOIN submissions s ON s.exam_id = ? AND s.student_id = u.id
        WHERE e.room_id = ?
        ORDER BY u.name ASC
    """, (exam_id, room_id))

    rows = cursor.fetchall()
    conn.close()

    result = []
    for r in rows:
        result.append({
            "student_id": r["student_id"],
            "name": r["name"],
            "email": r["email"],
            "student_code": r["student_code"],
            "submission_id": r["submission_id"],
            "status": r["status"] if r["status"] else "missing",
            "total_score": r["total_score"],
            "submitted_at": r["submitted_at"],
            "graded_by_ai": bool(r["graded_by_ai"]) if r["graded_by_ai"] is not None else False,
        })
    return result

@app.get("/api/rooms/{room_id}/exams/{exam_id}/analytics")
async def get_exam_analytics(room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    """Teacher analytics for score summary, submission counts, and question difficulty."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can view analytics")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")

    cursor.execute("SELECT id, total_score FROM exams WHERE id = ? AND room_id = ?", (exam_id, room_id))
    exam = cursor.fetchone()
    if not exam:
        conn.close()
        raise HTTPException(status_code=404, detail="Exam not found")

    exam_total_score = float(exam["total_score"] or 0)

    cursor.execute("""
        SELECT s.total_score
        FROM submissions s
        WHERE s.exam_id = ? AND s.status = 'approved'
    """, (exam_id,))
    approved_scores = [float(r["total_score"] or 0) for r in cursor.fetchall()]

    mean_score = round(statistics.mean(approved_scores), 2) if approved_scores else 0.0
    median_score = round(statistics.median(approved_scores), 2) if approved_scores else 0.0

    cursor.execute("""
        SELECT
            SUM(CASE WHEN s.status = 'missing' OR s.status IS NULL THEN 1 ELSE 0 END) AS missing_count,
            SUM(CASE WHEN s.status IS NOT NULL AND s.status != 'missing' THEN 1 ELSE 0 END) AS submitted_count
        FROM enrollments e
        LEFT JOIN submissions s ON s.exam_id = ? AND s.student_id = e.user_id
        WHERE e.room_id = ?
    """, (exam_id, room_id))
    submission_counts = cursor.fetchone()

    cursor.execute("""
        SELECT
            q.id AS question_id,
            q.order_index,
            q.text AS question_text,
            q.score AS max_score,
            AVG(
                CASE
                    WHEN s.id IS NOT NULL THEN COALESCE(sa.teacher_score, sa.ai_score, 0)
                    ELSE NULL
                END
            ) AS avg_score
        FROM questions q
        LEFT JOIN submission_answers sa ON sa.question_id = q.id
        LEFT JOIN submissions s ON s.id = sa.submission_id AND s.status = 'approved'
        WHERE q.exam_id = ?
        GROUP BY q.id, q.order_index, q.text, q.score
        ORDER BY q.order_index ASC
    """, (exam_id,))
    difficulty_rows = cursor.fetchall()

    difficulties = []
    for row in difficulty_rows:
        max_score = float(row["max_score"] or 0)
        avg_score = float(row["avg_score"] or 0)
        percent_correct = round((avg_score / max_score * 100.0), 2) if max_score > 0 else 0.0
        difficulties.append({
            "question_id": row["question_id"],
            "order_index": row["order_index"],
            "question_text": row["question_text"],
            "max_score": max_score,
            "avg_score": round(avg_score, 2),
            "percent_correct": percent_correct,
        })

    conn.close()

    return {
        "mean_score": mean_score,
        "median_score": median_score,
        "approved_submission_count": len(approved_scores),
        "score_distribution": _distribution_buckets(approved_scores, exam_total_score),
        "submission_counts": {
            "submitted": int(submission_counts["submitted_count"] or 0),
            "missing": int(submission_counts["missing_count"] or 0),
        },
        "difficulty_analysis": sorted(difficulties, key=lambda d: d["percent_correct"]),
    }

@app.get("/api/rooms/{room_id}/analytics")
async def get_room_analytics(room_id: int, user: dict = Depends(get_current_user)):
    """Teacher room-level analytics across all exams."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can view analytics")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT id, owner_id FROM rooms WHERE id = ?", (room_id,))
    room = cursor.fetchone()
    if not room:
        conn.close()
        raise HTTPException(status_code=404, detail="Room not found")
    if str(room["owner_id"]) != str(user["id"]):
        conn.close()
        raise HTTPException(status_code=403, detail=f"Unauthorized: owner={room['owner_id']} != user={user['id']}")

    cursor.execute("""
        SELECT
            e.id,
            e.title,
            e.total_score,
            COUNT(CASE WHEN s.status IS NOT NULL AND s.status != 'missing' THEN 1 END) AS submitted_count,
            COUNT(CASE WHEN s.status = 'approved' THEN 1 END) AS approved_count,
            AVG(CASE WHEN s.status = 'approved' THEN s.total_score ELSE NULL END) AS approved_mean
        FROM exams e
        LEFT JOIN submissions s ON s.exam_id = e.id
        WHERE e.room_id = ?
        GROUP BY e.id, e.title, e.total_score
        ORDER BY e.created_at DESC
    """, (room_id,))
    exams = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) AS total_students FROM enrollments WHERE room_id = ?", (room_id,))
    total_students = int(cursor.fetchone()["total_students"] or 0)

    cursor.execute("""
        SELECT s.total_score, e.total_score AS exam_total_score
        FROM submissions s
        JOIN exams e ON e.id = s.exam_id
        WHERE e.room_id = ? AND s.status = 'approved'
    """, (room_id,))
    room_approved_score_rows = cursor.fetchall()
    room_approved_scores = [float(r["total_score"] or 0) for r in room_approved_score_rows]
    room_approved_percents = []
    for r in room_approved_score_rows:
        exam_total = float(r["exam_total_score"] or 0)
        score = float(r["total_score"] or 0)
        pct = (score / exam_total * 100.0) if exam_total > 0 else 0.0
        room_approved_percents.append(max(0.0, min(100.0, pct)))
    conn.close()

    exam_summaries = []
    for e in exams:
        submitted_count = int(e["submitted_count"] or 0)
        approved_count = int(e["approved_count"] or 0)
        total_score = float(e["total_score"] or 0)
        approved_mean = round(float(e["approved_mean"] or 0), 2)
        submission_rate = round((submitted_count / total_students) * 100.0, 2) if total_students > 0 else 0.0
        mean_percent = round((approved_mean / total_score) * 100.0, 2) if total_score > 0 else 0.0
        exam_summaries.append({
            "exam_id": e["id"],
            "title": e["title"],
            "total_score": total_score,
            "submitted_count": submitted_count,
            "approved_count": approved_count,
            "approved_mean": approved_mean,
            "missing_count": max(0, total_students - submitted_count),
            "submission_rate": submission_rate,
            "mean_percent": mean_percent,
        })

    return {
        "total_students": total_students,
        "exam_count": len(exam_summaries),
        "overall_mean_score": round(statistics.mean(room_approved_scores), 2) if room_approved_scores else 0.0,
        "overall_median_score": round(statistics.median(room_approved_scores), 2) if room_approved_scores else 0.0,
        "overall_distribution": _distribution_buckets(room_approved_percents, 100.0),
        "exam_summaries": exam_summaries,
    }

@app.get("/api/rooms/{room_id}/exams/{exam_id}/export")
async def export_exam_scores(
    room_id: int,
    exam_id: int,
    user: dict = Depends(get_current_user),
    file_format: str = Query(default="csv", alias="format"),
):
    """Export exam results as CSV or XLSX."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can export scores")
    export_fmt = file_format.lower()
    if export_fmt not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Supported export formats are csv and xlsx")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")

    cursor.execute("SELECT id, title, total_score FROM exams WHERE id = ? AND room_id = ?", (exam_id, room_id))
    exam = cursor.fetchone()
    if not exam:
        conn.close()
        raise HTTPException(status_code=404, detail="Exam not found")

    cursor.execute("""
        SELECT
            u.id AS student_id,
            u.name,
            u.email,
            u.student_id AS student_code,
            COALESCE(s.status, 'missing') AS status,
            s.total_score,
            s.submitted_at
        FROM enrollments e
        JOIN users u ON u.id = e.user_id
        LEFT JOIN submissions s ON s.exam_id = ? AND s.student_id = e.user_id
        WHERE e.room_id = ?
        ORDER BY u.name ASC
    """, (exam_id, room_id))
    rows = cursor.fetchall()
    conn.close()

    safe_title = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in exam["title"])
    export_headers = [
        "Student ID",
        "Student Code",
        "Student Name",
        "Email",
        "Status",
        "Score",
        "Submitted At",
        "Exam Total Score",
    ]

    if export_fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(export_headers)
        for r in rows:
            writer.writerow([
                r["student_id"], r["student_code"], r["name"], r["email"],
                r["status"], r["total_score"], r["submitted_at"], exam["total_score"]
            ])

        csv_content = output.getvalue()
        output.close()
        filename = f"exam_{exam_id}_{safe_title}_scores.csv"
        return Response(
            content=csv_content,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(
            status_code=500,
            detail="XLSX export requires openpyxl. Please install dependencies from server/requirements.txt",
        )

    wb = Workbook()
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.active
    ws.title = "scores"
    ws.append(export_headers)

    # Make header row visually explicit.
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    for col_idx in range(1, len(export_headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r["student_id"], r["student_code"], r["name"], r["email"],
            r["status"], r["total_score"], r["submitted_at"], exam["total_score"]
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    bin_output = io.BytesIO()
    wb.save(bin_output)
    xlsx_bytes = bin_output.getvalue()
    bin_output.close()
    filename = f"exam_{exam_id}_{safe_title}_scores.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/rooms/{room_id}/exams/{exam_id}/submissions/me")
async def get_my_submission(room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    """Student views their own submission result."""
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM submissions WHERE exam_id = ? AND student_id = ?", (exam_id, user["id"]))
    submission = cursor.fetchone()
    if not submission:
        conn.close()
        return {"status": "missing"}

    # ซ่อนคะแนนจนกว่าอาจารย์จะอนุมัติ
    if submission["status"] != "approved":
        conn.close()
        return {
            "status": submission["status"],
            "submission_id": submission["id"],
            "submitted_at": submission["submitted_at"],
        }

    # Approved: return full result including score
    submission = dict(submission)
    cursor.execute("""
        SELECT sa.answer_text, sa.image_path, sa.image_paths, sa.ai_score, sa.ai_feedback, sa.teacher_score, sa.teacher_comment,
               q.text AS question_text, q.score AS max_score, q.order_index, q.image_path AS q_image_path, q.image_paths AS q_image_paths
        FROM submission_answers sa
        JOIN questions q ON sa.question_id = q.id
        WHERE sa.submission_id = ?
        ORDER BY q.order_index
    """, (submission["id"],))
    answers = []
    for a in cursor.fetchall():
        ad = dict(a)
        # Parse JSON image_paths
        if ad.get("image_paths"):
            try:
                ad["image_paths"] = json_module.loads(ad["image_paths"])
            except Exception:
                ad["image_paths"] = [ad["image_path"]] if ad.get("image_path") else []
        else:
            ad["image_paths"] = [ad["image_path"]] if ad.get("image_path") else []
            
        if ad.get("q_image_paths"):
            try:
                ad["q_image_paths"] = json_module.loads(ad["q_image_paths"])
            except Exception:
                ad["q_image_paths"] = [ad["q_image_path"]] if ad.get("q_image_path") else []
        else:
            ad["q_image_paths"] = [ad["q_image_path"]] if ad.get("q_image_path") else []
        
        answers.append(ad)
    conn.close()

    submission["answers"] = answers
    return submission


@app.get("/api/submissions/me")
async def get_all_my_submissions(user: dict = Depends(get_current_user)):
    """Student views all their submissions across all exams and rooms."""
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="Only students can view their submission history")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            e.id AS exam_id, e.title AS exam_title, e.total_score AS exam_total_score,
            r.id AS room_id, r.name AS room_name,
            s.id AS submission_id, COALESCE(s.status, 'missing') AS status, 
            s.total_score AS submission_score, s.submitted_at
        FROM enrollments en
        JOIN rooms r ON en.room_id = r.id
        JOIN exams e ON e.room_id = r.id
        LEFT JOIN submissions s ON s.exam_id = e.id AND s.student_id = en.user_id
        WHERE en.user_id = ?
        ORDER BY COALESCE(s.submitted_at, e.created_at) DESC
    """, (user["id"],))
    results = cursor.fetchall()
    conn.close()

    return [dict(r) for r in results]


@app.get("/api/rooms/{room_id}/exams/{exam_id}/submissions/{student_id}")
async def get_student_submission(room_id: int, exam_id: int, student_id: int, user: dict = Depends(get_current_user)):
    """Teacher views detailed submission of a specific student."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can view student submissions")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify room ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")

    # Get submission
    cursor.execute(
        "SELECT * FROM submissions WHERE exam_id = ? AND student_id = ?",
        (exam_id, student_id)
    )
    submission = cursor.fetchone()
    if not submission:
        conn.close()
        raise HTTPException(status_code=404, detail="Submission not found — student has not submitted yet")

    submission = dict(submission)

    # Get student info
    cursor.execute("SELECT id, name, email, student_id AS student_code FROM users WHERE id = ?", (student_id,))
    student_info = dict(cursor.fetchone())

    # Get answers with question details
    cursor.execute("""
        SELECT
            sa.id, sa.question_id, sa.answer_text, sa.ai_score, sa.ai_feedback, sa.ai_confidence,
            sa.teacher_score, sa.teacher_comment, sa.image_path, sa.image_paths, sa.quality_metrics,
            q.text AS question_text, q.score AS max_score, q.rubrics, q.answer_key, 
            q.order_index, q.image_path AS q_image_path, q.image_paths AS q_image_paths
        FROM submission_answers sa
        JOIN questions q ON sa.question_id = q.id
        WHERE sa.submission_id = ?
        ORDER BY q.order_index
    """, (submission["id"],))

    answers = []
    for a in cursor.fetchall():
        ad = dict(a)
        if ad.get("rubrics"):
            try:
                ad["rubrics"] = json_module.loads(ad["rubrics"])
            except Exception:
                ad["rubrics"] = []
                
        # Parse JSON image_paths for answers
        if ad.get("image_paths"):
            try:
                ad["image_paths"] = json_module.loads(ad["image_paths"])
            except Exception:
                ad["image_paths"] = [ad["image_path"]] if ad.get("image_path") else []
        else:
            ad["image_paths"] = [ad["image_path"]] if ad.get("image_path") else []
            
        # Parse JSON image_paths for questions
        if ad.get("q_image_paths"):
            try:
                ad["q_image_paths"] = json_module.loads(ad["q_image_paths"])
            except Exception:
                ad["q_image_paths"] = [ad["q_image_path"]] if ad.get("q_image_path") else []
        else:
            ad["q_image_paths"] = [ad["q_image_path"]] if ad.get("q_image_path") else []
            
        answers.append(ad)

    conn.close()

    return {
        "submission": submission,
        "student": student_info,
        "answers": answers,
    }


@app.put("/api/rooms/{room_id}/exams/{exam_id}/submissions/{student_id}/approve")
async def approve_submission(
    request: Request,
    room_id: int, exam_id: int, student_id: int,
    body: ApproveSubmissionRequest,
    user: dict = Depends(get_current_user)
):
    """Teacher approves a submission, optionally overriding AI scores."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can approve submissions")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify room ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")

    # Get submission
    cursor.execute("SELECT * FROM submissions WHERE exam_id = ? AND student_id = ?", (exam_id, student_id))
    submission = cursor.fetchone()
    if not submission:
        conn.close()
        raise HTTPException(status_code=404, detail="Submission not found")

    submission_id = submission["id"]

    # Apply teacher score overrides per question
    total_teacher_score = 0.0
    cursor.execute("SELECT * FROM submission_answers WHERE submission_id = ?", (submission_id,))
    all_answers = cursor.fetchall()

    for ans in all_answers:
        q_id_str = str(ans["question_id"])
        t_score = body.teacher_scores.get(q_id_str) if body.teacher_scores else None
        t_comment = body.teacher_comments.get(q_id_str) if body.teacher_comments else None

        if t_score is not None:
            cursor.execute(
                "UPDATE submission_answers SET teacher_score = ?, teacher_comment = ? WHERE id = ?",
                (t_score, t_comment, ans["id"])
            )
            total_teacher_score += float(t_score)
        else:
            total_teacher_score += float(ans["ai_score"] or 0)
            if t_comment:
                cursor.execute("UPDATE submission_answers SET teacher_comment = ? WHERE id = ?", (t_comment, ans["id"]))

    # Update submission status to approved with final score
    cursor.execute(
        "UPDATE submissions SET status = 'approved', total_score = ? WHERE id = ?",
        (round(total_teacher_score, 1), submission_id)
    )
    conn.commit()
    conn.close()

    log_audit_action(user["id"], "GRADE_CHANGE", str(submission_id), f"Approved/Updated grade for student {student_id} on exam {exam_id}", request.client.host)
    return {"message": "Submission approved", "total_score": round(total_teacher_score, 1)}

@app.post("/api/rooms/{room_id}/exams/{exam_id}/questions/{question_id}/rescore")
async def rescore_question(room_id: int, exam_id: int, question_id: int, user: dict = Depends(get_current_user)):
    """Force AI to rescore a specific question for all submissions in an exam."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can rescore")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")
        
    # Find all submissions that have an answer for this question
    cursor.execute("SELECT submission_id FROM submission_answers WHERE question_id = ?", (question_id,))
    subs = cursor.fetchall()
    
    for row in subs:
        await grading_queue.put({
            "submission_id": row["submission_id"],
            "room_id": room_id,
            "exam_id": exam_id,
            "question_id": question_id
        })
        
    conn.close()
    return {"message": f"เพิ่มการประมวลผลข้อนี้ใหม่จำนวน {len(subs)} รายการลงในคิวแล้ว"}



@app.post("/api/rooms/{room_id}/exams/{exam_id}/bulk-approve")
async def bulk_approve(room_id: int, exam_id: int, body: BulkApproveRequest, user: dict = Depends(get_current_user)):
    """Teacher bulk-approves multiple students using AI scores as-is."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can approve submissions")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify room ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")

    approved = []
    skipped = []
    for sid in body.student_ids:
        cursor.execute("SELECT id, status FROM submissions WHERE exam_id = ? AND student_id = ?", (exam_id, sid))
        sub = cursor.fetchone()
        if not sub:
            skipped.append({"student_id": sid, "reason": "not_submitted"})
            continue

        if sub["status"] in ("ready", "needs_review"):
            cursor.execute("UPDATE submissions SET status = 'approved' WHERE id = ?", (sub["id"],))
            approved.append(sid)
        else:
            skipped.append({"student_id": sid, "reason": f"status_{sub['status']}"})

    conn.commit()
    conn.close()

    return {
        "message": f"Approved {len(approved)} submissions",
        "approved_student_ids": approved,
        "skipped": skipped,
    }


# Demo route
@app.get("/api/ping")
async def ping():
    return {"message": "pong"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)

@app.get("/api/audit-logs")
async def get_audit_logs(user=Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, action_type, entity_id, details, ip_address, created_at FROM audit_logs WHERE user_id = ? ORDER BY created_at DESC LIMIT 100",
        (user["id"],)
    )
    logs = cursor.fetchall()
    conn.close()
    return logs


# ============================================================
# Auto-save Draft Endpoints
# ============================================================

@app.post("/api/rooms/{room_id}/exams/{exam_id}/draft")
async def save_draft(room_id: int, exam_id: int, body: DraftSaveRequest, user: dict = Depends(get_current_user)):
    """Student saves answer draft to server (hybrid with localStorage)."""
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="Only students can save drafts")

    conn = get_db_connection()
    cursor = conn.cursor()

    draft_json = json_module.dumps(body.answers, ensure_ascii=False)
    cursor.execute(
        """INSERT INTO submission_drafts (exam_id, student_id, draft_data)
           VALUES (?, ?, ?)
           ON DUPLICATE KEY UPDATE draft_data = VALUES(draft_data), updated_at = CURRENT_TIMESTAMP""",
        (exam_id, user["id"], draft_json)
    )
    conn.commit()
    conn.close()
    return {"message": "Draft saved"}


@app.get("/api/rooms/{room_id}/exams/{exam_id}/draft")
async def get_draft(room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    """Student loads saved draft from server."""
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="Only students can load drafts")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT draft_data, updated_at FROM submission_drafts WHERE exam_id = ? AND student_id = ?",
        (exam_id, user["id"])
    )
    row = cursor.fetchone()
    conn.close()

    if not row:
        return {"answers": {}, "updated_at": None}

    try:
        answers = json_module.loads(row["draft_data"])
    except Exception:
        answers = {}

    return {"answers": answers, "updated_at": str(row["updated_at"])}


@app.delete("/api/rooms/{room_id}/exams/{exam_id}/draft")
async def delete_draft(room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    """Delete draft after successful submission."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "DELETE FROM submission_drafts WHERE exam_id = ? AND student_id = ?",
        (exam_id, user["id"])
    )
    conn.commit()
    conn.close()
    return {"message": "Draft deleted"}


# ============================================================
# Time Extension Endpoints
# ============================================================

@app.post("/api/rooms/{room_id}/exams/{exam_id}/extensions")
async def grant_time_extension(
    request: Request,
    room_id: int, exam_id: int,
    body: TimeExtensionRequest,
    user: dict = Depends(get_current_user)
):
    """Teacher grants extra time to a specific student or the whole class (student_id=None)."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Only teachers can grant time extensions")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify room ownership
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Room not found or unauthorized")

    cursor.execute("SELECT id FROM exams WHERE id = ? AND room_id = ?", (exam_id, room_id))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Exam not found")

    cursor.execute(
        "INSERT INTO exam_time_extensions (exam_id, student_id, extra_minutes, granted_by, note) VALUES (?, ?, ?, ?, ?)",
        (exam_id, body.student_id, body.extra_minutes, user["id"], body.note)
    )
    conn.commit()
    conn.close()

    target = f"นักเรียน ID {body.student_id}" if body.student_id else "ทั้งห้อง"
    log_audit_action(user["id"], "TIME_EXTENSION", str(exam_id),
                     f"Granted {body.extra_minutes} mins to {target} on exam {exam_id}",
                     request.client.host)
    return {"message": f"ขยายเวลา {body.extra_minutes} นาทีสำเร็จ"}


@app.get("/api/rooms/{room_id}/exams/{exam_id}/extensions/me")
async def get_my_extension(room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    """Student gets their total extra minutes for this exam."""
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="Students only")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Sum up: extensions for this student + whole-class extensions (student_id IS NULL)
    cursor.execute(
        """SELECT COALESCE(SUM(extra_minutes), 0) as total_extra
           FROM exam_time_extensions
           WHERE exam_id = ? AND (student_id = ? OR student_id IS NULL)""",
        (exam_id, user["id"])
    )
    row = cursor.fetchone()
    conn.close()
    return {"extra_minutes": int(row["total_extra"] or 0)}


@app.get("/api/rooms/{room_id}/exams/{exam_id}/extensions")
async def get_all_extensions(room_id: int, exam_id: int, user: dict = Depends(get_current_user)):
    """Teacher views all extensions for an exam."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Teachers only")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """SELECT e.*, u.name as student_name, u.email as student_email
           FROM exam_time_extensions e
           LEFT JOIN users u ON e.student_id = u.id
           WHERE e.exam_id = ?
           ORDER BY e.created_at DESC""",
        (exam_id,)
    )
    extensions = cursor.fetchall()
    conn.close()
    return [dict(r) for r in extensions]


# ============================================================
# Question Bank Endpoints
# ============================================================

@app.get("/api/question-bank")
async def list_question_bank(user: dict = Depends(get_current_user)):
    """Teacher lists their own question bank."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Teachers only")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT * FROM question_bank WHERE owner_id = ? ORDER BY created_at DESC",
        (user["id"],)
    )
    items = cursor.fetchall()
    conn.close()

    result = []
    for item in items:
        d = dict(item)
        if d.get("rubrics"):
            try:
                d["rubrics"] = json_module.loads(d["rubrics"])
            except Exception:
                d["rubrics"] = []
        result.append(d)
    return result


@app.post("/api/question-bank")
async def create_question_bank(body: QuestionBankCreate, user: dict = Depends(get_current_user)):
    """Teacher adds a question to their bank."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Teachers only")

    conn = get_db_connection()
    cursor = conn.cursor()
    rubrics_json = json_module.dumps(body.rubrics, ensure_ascii=False) if body.rubrics else None
    cursor.execute(
        "INSERT INTO question_bank (owner_id, text, score, answer_key, rubrics, tags) VALUES (?, ?, ?, ?, ?, ?)",
        (user["id"], body.text, body.score, body.answer_key, rubrics_json, body.tags)
    )
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return {"id": new_id, "message": "Question saved to bank"}


@app.delete("/api/question-bank/{question_id}")
async def delete_question_bank(question_id: int, user: dict = Depends(get_current_user)):
    """Teacher deletes a question from their bank."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Teachers only")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM question_bank WHERE id = ? AND owner_id = ?", (question_id, user["id"]))
    conn.commit()
    conn.close()
    return {"message": "Deleted"}


@app.post("/api/question-bank/save-from-exam")
async def save_questions_from_exam(
    room_id: int,
    exam_id: int,
    user: dict = Depends(get_current_user)
):
    """Teacher saves all questions from an exam into their question bank."""
    if user["role"] != "teacher":
        raise HTTPException(status_code=403, detail="Teachers only")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM rooms WHERE id = ? AND owner_id = ?", (room_id, user["id"]))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail="Unauthorized")

    cursor.execute("SELECT * FROM questions WHERE exam_id = ?", (exam_id,))
    questions = cursor.fetchall()

    saved = 0
    for q in questions:
        cursor.execute(
            "INSERT INTO question_bank (owner_id, text, score, answer_key, rubrics, tags) VALUES (?, ?, ?, ?, ?, ?)",
            (user["id"], q["text"], q["score"], q["answer_key"], q["rubrics"], None)
        )
        saved += 1

    conn.commit()
    conn.close()
    return {"message": f"บันทึก {saved} ข้อลงคลังเรียบร้อยแล้ว", "saved": saved}


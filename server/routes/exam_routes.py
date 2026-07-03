from fastapi import APIRouter, Depends, HTTPException, status, Header, UploadFile, File, Form, Request, Query, BackgroundTasks
from fastapi.responses import Response, StreamingResponse
import pymysql
import json
import csv
import io
import time
import asyncio
import statistics
from typing import Optional, List

from server.database import get_db_connection
from server.auth import get_password_hash, verify_password, create_access_token, decode_token
from server.models import *
from server.utils import check_rate_limit, upload_to_cloudinary, get_current_user, grading_queue, trigger_socket_notify, _distribution_buckets
from server.services.ai_service import _USE_GEMINI, _genai_client, _GEMINI_MODEL

router = APIRouter(prefix="/api/rooms/{room_id}/exams", tags=["Exams"])

@router.post('')
async def create_exam(room_id: int, exam: ExamCreate, user: dict=Depends(get_current_user)):
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can create exams')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail='Room not found or unauthorized')
    computed_total_score = sum((float(q.score) for q in exam.questions))
    cursor.execute('INSERT INTO exams (room_id, title, description, total_score, start_date, end_date, is_randomized) VALUES (?, ?, ?, ?, ?, ?, ?)', (room_id, exam.title, exam.description, computed_total_score, exam.start_date, exam.end_date, exam.is_randomized))
    exam_id = cursor.lastrowid
    import base64, re as _re
    for q in exam.questions:
        rubrics_json = json.dumps(q.rubrics, ensure_ascii=False) if q.rubrics else None
        image_paths = []
        import time as _time
        for img_idx, img_b64 in enumerate(q.question_images_base64 or []):
            try:
                match = _re.match('data:(?P<mime>[^;]+);base64,(?P<data>.+)', img_b64)
                if match:
                    mime = match.group('mime')
                    b64data = match.group('data')
                    img_bytes = base64.b64decode(b64data)
                    ext = mime.split('/')[-1].replace('jpeg', 'jpg')
                    c_url = upload_to_cloudinary(img_bytes, folder=f'questions/{exam_id}')
                    if c_url:
                        image_paths.append(c_url)
            except Exception as img_err:
                print(f'[Question Image] Failed to save index {img_idx}: {img_err}')
        image_paths_json = json.dumps(image_paths) if image_paths else None
        first_image = image_paths[0] if image_paths else None
        cursor.execute('INSERT INTO questions (exam_id, text, score, answer_key, rubrics, order_index, image_path, image_paths) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', (exam_id, q.text, q.score, q.answer_key, rubrics_json, q.order_index, first_image, image_paths_json))
    conn.commit()
    cursor.execute('SELECT * FROM exams WHERE id = ?', (exam_id,))
    new_exam = dict(cursor.fetchone())
    conn.close()
    return new_exam

@router.put('/{exam_id}')
async def update_exam(room_id: int, exam_id: int, exam: ExamCreate, user: dict=Depends(get_current_user)):
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can update exams')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM exams WHERE id = ? AND room_id = ?', (exam_id, room_id))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail='Exam not found')
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail='Room not found or unauthorized')
    
    # Check if there are any submissions before allowing edits
    cursor.execute('SELECT id FROM submissions WHERE exam_id = ?', (exam_id,))
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail='ไม่สามารถแก้ไขข้อสอบได้ เนื่องจากมีนักเรียนเริ่มทำหรือส่งคำตอบมาแล้ว')
        
    computed_total_score = sum((float(q.score) for q in exam.questions))
    cursor.execute('UPDATE exams SET title = ?, description = ?, total_score = ?, start_date = ?, end_date = ?, is_randomized = ? WHERE id = ?', (exam.title, exam.description, computed_total_score, exam.start_date, exam.end_date, exam.is_randomized, exam_id))
    cursor.execute('DELETE FROM questions WHERE exam_id = ?', (exam_id,))
    import base64, re as _re
    for q in exam.questions:
        rubrics_json = json.dumps(q.rubrics, ensure_ascii=False) if q.rubrics else None
        image_paths = []
        import time as _time
        for img_idx, img_b64 in enumerate(q.question_images_base64 or []):
            try:
                match = _re.match('data:(?P<mime>[^;]+);base64,(?P<data>.+)', img_b64)
                if match:
                    mime = match.group('mime')
                    b64data = match.group('data')
                    img_bytes = base64.b64decode(b64data)
                    ext = mime.split('/')[-1].replace('jpeg', 'jpg')
                    c_url = upload_to_cloudinary(img_bytes, folder=f'questions/{exam_id}')
                    if c_url:
                        image_paths.append(c_url)
            except Exception as img_err:
                print(f'[Question Image] Failed to save index {img_idx}: {img_err}')
        image_paths_json = json.dumps(image_paths) if image_paths else None
        first_image = image_paths[0] if image_paths else None
        cursor.execute('INSERT INTO questions (exam_id, text, score, answer_key, rubrics, order_index, image_path, image_paths) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', (exam_id, q.text, q.score, q.answer_key, rubrics_json, q.order_index, first_image, image_paths_json))
    conn.commit()
    cursor.execute('SELECT * FROM exams WHERE id = ?', (exam_id,))
    updated_exam = dict(cursor.fetchone())
    conn.close()
    return updated_exam

@router.delete('/{exam_id}')
async def delete_exam(request: Request, room_id: int, exam_id: int, user: dict=Depends(get_current_user)):
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can delete exams')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail='Room not found or unauthorized')
    cursor.execute('SELECT id FROM exams WHERE id = ? AND room_id = ?', (exam_id, room_id))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail='Exam not found')
    cursor.execute('DELETE FROM exams WHERE id = ?', (exam_id,))
    conn.commit()
    conn.close()
    return {'message': 'Exam deleted successfully'}

@router.get('')
async def list_exams(room_id: int, user: dict=Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()
    if user['role'] == 'teacher':
        cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    else:
        cursor.execute('SELECT room_id FROM enrollments WHERE room_id = ? AND user_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail='Unauthorized')
    cursor.execute('SELECT * FROM exams WHERE room_id = ? ORDER BY created_at DESC', (room_id,))
    exams = cursor.fetchall()
    conn.close()
    return [dict(e) for e in exams]

@router.get('/{exam_id}')
async def get_exam(room_id: int, exam_id: int, user: dict=Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM exams WHERE id = ? AND room_id = ?', (exam_id, room_id))
    exam = cursor.fetchone()
    if not exam:
        conn.close()
        raise HTTPException(status_code=404, detail='Exam not found')
    cursor.execute('SELECT * FROM questions WHERE exam_id = ? ORDER BY order_index', (exam_id,))
    questions = cursor.fetchall()
    conn.close()
    result = dict(exam)
    result['questions'] = []
    for q in questions:
        qd = dict(q)
        if qd.get('rubrics'):
            try:
                qd['rubrics'] = json.loads(qd['rubrics'])
            except Exception:
                qd['rubrics'] = []
        if qd.get('image_paths'):
            try:
                qd['image_paths'] = json.loads(qd['image_paths'])
            except Exception:
                qd['image_paths'] = [qd.get('image_path')] if qd.get('image_path') else []
        else:
            qd['image_paths'] = [qd['image_path']] if qd.get('image_path') else []
        result['questions'].append(qd)
    if result.get('is_randomized') and user['role'] == 'student':
        import random as _random
        rng = _random.Random(user['id'] + exam_id * 1000)
        rng.shuffle(result['questions'])
    return result

@router.post('/{exam_id}/submit')
async def submit_exam(room_id: int, exam_id: int, answers: str=Form(...), user: dict=Depends(get_current_user), **kwargs):
    """Student submits answers (lock-once). Supports text + image per question."""
    raise HTTPException(status_code=500, detail='Use the multipart endpoint below')

@router.post('/{exam_id}/submit-multipart')
async def submit_exam_multipart(request: Request, room_id: int, exam_id: int, user: dict=Depends(get_current_user)):
    """Student submits answers with optional image per question (lock-once, multipart/form-data)."""
    if user['role'] != 'student':
        raise HTTPException(status_code=403, detail='Only students can submit exams')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM exams WHERE id = ? AND room_id = ?', (exam_id, room_id))
    exam_row = cursor.fetchone()
    if not exam_row:
        conn.close()
        raise HTTPException(status_code=404, detail='Exam not found')
    end_date_str = exam_row['end_date']
    if end_date_str:
        try:
            from datetime import datetime, timezone, timedelta
            deadline = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
            if deadline.tzinfo is None:
                deadline = deadline.replace(tzinfo=timezone.utc)
                
            now = datetime.now(timezone.utc)
            if now > deadline + timedelta(minutes=1): # Add 1 min grace period
                conn.close()
                raise HTTPException(status_code=403, detail='เลยกำหนดเวลาส่งคำตอบข้อสอบแล้ว (Deadline Passed)')
        except (ValueError, TypeError):
            pass
            
    cursor.execute('SELECT id, status FROM submissions WHERE exam_id = ? AND student_id = ?', (exam_id, user['id']))
    existing = cursor.fetchone()
    if existing and existing['status'] != 'missing':
        conn.close()
        raise HTTPException(status_code=409, detail='คุณได้ส่งคำตอบข้อสอบนี้ไปแล้ว ไม่สามารถส่งซ้ำได้')
    form = await request.form()
    answers_json = form.get('answers', '[]')
    try:
        answers_list = json.loads(answers_json)
    except Exception:
        conn.close()
        raise HTTPException(status_code=400, detail='Invalid answers JSON')
    answers_map = {int(a['question_id']): a.get('answer_text', '') for a in answers_list}
    cursor.execute("INSERT INTO submissions (exam_id, student_id, status, submitted_at, graded_by_ai)\n           VALUES (?, ?, 'submitted', CURRENT_TIMESTAMP, 0)\n           ON DUPLICATE KEY UPDATE\n             status='submitted', submitted_at=CURRENT_TIMESTAMP, graded_by_ai=0", (exam_id, user['id']))
    submission_id = cursor.lastrowid
    if not submission_id:
        cursor.execute('SELECT id FROM submissions WHERE exam_id = ? AND student_id = ?', (exam_id, user['id']))
        submission_id = cursor.fetchone()['id']
    cursor.execute('SELECT * FROM questions WHERE exam_id = ? ORDER BY order_index', (exam_id,))
    questions = {q['id']: dict(q) for q in cursor.fetchall()}
    total_ai_score = 0.0
    for q_id, q in questions.items():
        base_answer_text = answers_map.get(q_id, '')
        answer_text = base_answer_text[:300] if base_answer_text else ''
        img_list: list[bytes] = []
        img_mime_list: list[str] = []
        img_paths: list[str] = []

        for img_idx in range(10):
            field_name = f'image_{q_id}_{img_idx}'
            file_field = form.get(field_name)
            if not file_field and img_idx == 0:
                file_field = form.get(f'image_{q_id}')
            if not file_field or not hasattr(file_field, 'read'):
                continue
            raw_bytes = await file_field.read()
            if not raw_bytes:
                continue
            mime = file_field.content_type or 'image/jpeg'
            ext = mime.split('/')[-1].replace('jpeg', 'jpg')
            c_url = upload_to_cloudinary(raw_bytes, folder=f"submissions/{exam_id}/{user['id']}")
            if c_url:
                img_paths.append(c_url)
        image_paths_json = json.dumps(img_paths) if img_paths else None
        first_image_path = img_paths[0] if img_paths else None
        cursor.execute("INSERT INTO submission_answers\n                 (submission_id, question_id, answer_text, ai_score, ai_feedback, ai_confidence, image_path, image_paths)\n               VALUES (?, ?, ?, 0, '', 'medium', ?, ?)\n               ON DUPLICATE KEY UPDATE\n                 answer_text=VALUES(answer_text), image_path=VALUES(image_path), image_paths=VALUES(image_paths)", (submission_id, q_id, answer_text, first_image_path, image_paths_json))
    await grading_queue.put({'submission_id': submission_id, 'room_id': room_id, 'exam_id': exam_id, 'user_id': user['id']})
    conn.commit()
    conn.close()
    return {'message': 'ส่งคำตอบสำเร็จ (ระบบกำลังตรวจคะแนน)', 'submission_id': submission_id, 'status': 'submitted', 'ai_score': 0}

@router.get('/{exam_id}/export-csv')
async def export_exam_csv(room_id: int, exam_id: int, user: dict=Depends(get_current_user)):
    """Teacher exports exam scores to CSV"""
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can export scores')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail='Unauthorized')
    cursor.execute("\n        SELECT \n            u.student_id AS student_code, u.name, u.email, \n            COALESCE(s.status, 'missing') as status, \n            s.total_score, s.submitted_at\n        FROM enrollments e\n        JOIN users u ON e.user_id = u.id\n        LEFT JOIN submissions s ON s.exam_id = ? AND s.student_id = u.id\n        WHERE e.room_id = ?\n        ORDER BY u.name ASC\n    ", (exam_id, room_id))
    results = cursor.fetchall()
    cursor.execute('SELECT title FROM exams WHERE id = ?', (exam_id,))
    row = cursor.fetchone()
    exam_title = row['title'] if row else 'Exam'
    conn.close()
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(['Student ID', 'Name', 'Status', 'Score', 'Submitted At'])
    for r in results:
        writer.writerow([r['student_code'] or '-', r['name'], r['status'], r['total_score'] if r['total_score'] is not None else '0', r['submitted_at'] or '-'])
    content = output.getvalue()
    return StreamingResponse(iter([content]), media_type='text/csv', headers={'Content-Disposition': f"attachment; filename=scores_{exam_title.replace(' ', '_')}.csv"})

@router.get('/{exam_id}/submissions')
async def list_submissions(room_id: int, exam_id: int, user: dict=Depends(get_current_user)):
    """Teacher gets list of all enrolled students with their submission status."""
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can view all submissions')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail='Unauthorized')
    cursor.execute('\n        SELECT\n            u.id AS student_id, u.name, u.email, u.student_id AS student_code,\n            s.id AS submission_id, s.status, s.total_score, s.submitted_at, s.graded_by_ai\n        FROM enrollments e\n        JOIN users u ON e.user_id = u.id\n        LEFT JOIN submissions s ON s.exam_id = ? AND s.student_id = u.id\n        WHERE e.room_id = ?\n        ORDER BY u.name ASC\n    ', (exam_id, room_id))
    rows = cursor.fetchall()
    conn.close()
    result = []
    for r in rows:
        result.append({'student_id': r['student_id'], 'name': r['name'], 'email': r['email'], 'student_code': r['student_code'], 'submission_id': r['submission_id'], 'status': r['status'] if r['status'] else 'missing', 'total_score': r['total_score'], 'submitted_at': r['submitted_at'], 'graded_by_ai': bool(r['graded_by_ai']) if r['graded_by_ai'] is not None else False})
    return result

@router.get('/{exam_id}/analytics')
async def get_exam_analytics(room_id: int, exam_id: int, user: dict=Depends(get_current_user)):
    """Teacher analytics for score summary, submission counts, and question difficulty."""
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can view analytics')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail='Unauthorized')
    cursor.execute('SELECT id, total_score FROM exams WHERE id = ? AND room_id = ?', (exam_id, room_id))
    exam = cursor.fetchone()
    if not exam:
        conn.close()
        raise HTTPException(status_code=404, detail='Exam not found')
    exam_total_score = float(exam['total_score'] or 0)
    cursor.execute("\n        SELECT s.total_score\n        FROM submissions s\n        WHERE s.exam_id = ? AND s.status = 'approved'\n    ", (exam_id,))
    approved_scores = [float(r['total_score'] or 0) for r in cursor.fetchall()]
    mean_score = round(statistics.mean(approved_scores), 2) if approved_scores else 0.0
    median_score = round(statistics.median(approved_scores), 2) if approved_scores else 0.0
    cursor.execute("\n        SELECT\n            SUM(CASE WHEN s.status = 'missing' OR s.status IS NULL THEN 1 ELSE 0 END) AS missing_count,\n            SUM(CASE WHEN s.status IS NOT NULL AND s.status != 'missing' THEN 1 ELSE 0 END) AS submitted_count\n        FROM enrollments e\n        LEFT JOIN submissions s ON s.exam_id = ? AND s.student_id = e.user_id\n        WHERE e.room_id = ?\n    ", (exam_id, room_id))
    submission_counts = cursor.fetchone()
    cursor.execute("\n        SELECT\n            q.id AS question_id,\n            q.order_index,\n            q.text AS question_text,\n            q.score AS max_score,\n            AVG(\n                CASE\n                    WHEN s.id IS NOT NULL THEN COALESCE(sa.teacher_score, sa.ai_score, 0)\n                    ELSE NULL\n                END\n            ) AS avg_score\n        FROM questions q\n        LEFT JOIN submission_answers sa ON sa.question_id = q.id\n        LEFT JOIN submissions s ON s.id = sa.submission_id AND s.status = 'approved'\n        WHERE q.exam_id = ?\n        GROUP BY q.id, q.order_index, q.text, q.score\n        ORDER BY q.order_index ASC\n    ", (exam_id,))
    difficulty_rows = cursor.fetchall()
    difficulties = []
    for row in difficulty_rows:
        max_score = float(row['max_score'] or 0)
        avg_score = float(row['avg_score'] or 0)
        percent_correct = round(avg_score / max_score * 100.0, 2) if max_score > 0 else 0.0
        difficulties.append({'question_id': row['question_id'], 'order_index': row['order_index'], 'question_text': row['question_text'], 'max_score': max_score, 'avg_score': round(avg_score, 2), 'percent_correct': percent_correct})
    conn.close()
    return {'mean_score': mean_score, 'median_score': median_score, 'approved_submission_count': len(approved_scores), 'score_distribution': _distribution_buckets(approved_scores, exam_total_score), 'submission_counts': {'submitted': int(submission_counts['submitted_count'] or 0), 'missing': int(submission_counts['missing_count'] or 0)}, 'difficulty_analysis': sorted(difficulties, key=lambda d: d['percent_correct'])}

@router.get('/{exam_id}/export')
async def export_exam_scores(room_id: int, exam_id: int, user: dict=Depends(get_current_user), file_format: str=Query(default='csv', alias='format')):
    """Export exam results as CSV or XLSX."""
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can export scores')
    export_fmt = file_format.lower()
    if export_fmt not in ('csv', 'xlsx'):
        raise HTTPException(status_code=400, detail='Supported export formats are csv and xlsx')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail='Unauthorized')
    cursor.execute('SELECT id, title, total_score FROM exams WHERE id = ? AND room_id = ?', (exam_id, room_id))
    exam = cursor.fetchone()
    if not exam:
        conn.close()
        raise HTTPException(status_code=404, detail='Exam not found')
    cursor.execute("\n        SELECT\n            u.id AS student_id,\n            u.name,\n            u.email,\n            u.student_id AS student_code,\n            COALESCE(s.status, 'missing') AS status,\n            s.total_score,\n            s.submitted_at\n        FROM enrollments e\n        JOIN users u ON u.id = e.user_id\n        LEFT JOIN submissions s ON s.exam_id = ? AND s.student_id = e.user_id\n        WHERE e.room_id = ?\n        ORDER BY u.name ASC\n    ", (exam_id, room_id))
    rows = cursor.fetchall()
    conn.close()
    safe_title = ''.join((ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in exam['title']))
    export_headers = ['รหัสนักศึกษา', 'ชื่อ-นามสกุล', 'สถานะ', 'คะแนนที่ได้', 'คะแนนเต็ม', 'วันที่ส่ง']
    if export_fmt == 'csv':
        output = io.StringIO()
        output.write('\ufeff')
        writer = csv.writer(output)
        writer.writerow(export_headers)
        for r in rows:
            writer.writerow([r['student_code'], r['name'], r['status'], r['total_score'], exam['total_score'], r['submitted_at']])
        csv_content = output.getvalue()
        output.close()
        filename = f'exam_{exam_id}_{safe_title}_scores.csv'
        return Response(content=csv_content, media_type='text/csv; charset=utf-8', headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=500, detail='XLSX export requires openpyxl. Please install dependencies from server/requirements.txt')
    wb = Workbook()
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.active
    ws.title = 'scores'
    ws.append(export_headers)
    header_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    for col_idx in range(1, len(export_headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for r in rows:
        ws.append([r['student_code'], r['name'], r['status'], r['total_score'], exam['total_score'], r['submitted_at']])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    bin_output = io.BytesIO()
    wb.save(bin_output)
    xlsx_bytes = bin_output.getvalue()
    bin_output.close()
    filename = f'exam_{exam_id}_{safe_title}_scores.xlsx'
    return Response(content=xlsx_bytes, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="{filename}"'})

@router.get('/{exam_id}/submissions/me')
async def get_my_submission(room_id: int, exam_id: int, user: dict=Depends(get_current_user)):
    """Student views their own submission result."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM submissions WHERE exam_id = ? AND student_id = ?', (exam_id, user['id']))
    submission = cursor.fetchone()
    if not submission:
        conn.close()
        return {'status': 'missing'}
    if submission['status'] != 'approved':
        conn.close()
        return {'status': submission['status'], 'submission_id': submission['id'], 'submitted_at': submission['submitted_at']}
    submission = dict(submission)
    cursor.execute('\n        SELECT sa.answer_text, sa.image_path, sa.image_paths, sa.ai_score, sa.ai_feedback, sa.teacher_score, sa.teacher_comment,\n               q.text AS question_text, q.score AS max_score, q.order_index, q.image_path AS q_image_path, q.image_paths AS q_image_paths\n        FROM submission_answers sa\n        JOIN questions q ON sa.question_id = q.id\n        WHERE sa.submission_id = ?\n        ORDER BY q.order_index\n    ', (submission['id'],))
    answers = []
    for a in cursor.fetchall():
        ad = dict(a)
        if ad.get('image_paths'):
            try:
                ad['image_paths'] = json.loads(ad['image_paths'])
            except Exception:
                ad['image_paths'] = [ad['image_path']] if ad.get('image_path') else []
        else:
            ad['image_paths'] = [ad['image_path']] if ad.get('image_path') else []
        if ad.get('q_image_paths'):
            try:
                ad['q_image_paths'] = json.loads(ad['q_image_paths'])
            except Exception:
                ad['q_image_paths'] = [ad['q_image_path']] if ad.get('q_image_path') else []
        else:
            ad['q_image_paths'] = [ad['q_image_path']] if ad.get('q_image_path') else []
        answers.append(ad)
    conn.close()
    submission['answers'] = answers
    return submission

@router.get('/{exam_id}/submissions/{student_id}')
async def get_student_submission(room_id: int, exam_id: int, student_id: int, user: dict=Depends(get_current_user)):
    """Teacher views detailed submission of a specific student."""
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can view student submissions')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail='Unauthorized')
    cursor.execute('SELECT * FROM submissions WHERE exam_id = ? AND student_id = ?', (exam_id, student_id))
    submission = cursor.fetchone()
    if not submission:
        conn.close()
        raise HTTPException(status_code=404, detail='Submission not found — student has not submitted yet')
    submission = dict(submission)
    cursor.execute('SELECT id, name, email, student_id AS student_code FROM users WHERE id = ?', (student_id,))
    student_info = dict(cursor.fetchone())
    cursor.execute('\n        SELECT\n            sa.id, sa.question_id, sa.answer_text, sa.ai_score, sa.ai_feedback, sa.ai_confidence,\n            sa.teacher_score, sa.teacher_comment, sa.image_path, sa.image_paths, sa.quality_metrics,\n            q.text AS question_text, q.score AS max_score, q.rubrics, q.answer_key, \n            q.order_index, q.image_path AS q_image_path, q.image_paths AS q_image_paths\n        FROM submission_answers sa\n        JOIN questions q ON sa.question_id = q.id\n        WHERE sa.submission_id = ?\n        ORDER BY q.order_index\n    ', (submission['id'],))
    answers = []
    for a in cursor.fetchall():
        ad = dict(a)
        if ad.get('rubrics'):
            try:
                ad['rubrics'] = json.loads(ad['rubrics'])
            except Exception:
                ad['rubrics'] = []
        if ad.get('image_paths'):
            try:
                ad['image_paths'] = json.loads(ad['image_paths'])
            except Exception:
                ad['image_paths'] = [ad['image_path']] if ad.get('image_path') else []
        else:
            ad['image_paths'] = [ad['image_path']] if ad.get('image_path') else []
        if ad.get('q_image_paths'):
            try:
                ad['q_image_paths'] = json.loads(ad['q_image_paths'])
            except Exception:
                ad['q_image_paths'] = [ad['q_image_path']] if ad.get('q_image_path') else []
        else:
            ad['q_image_paths'] = [ad['q_image_path']] if ad.get('q_image_path') else []
        answers.append(ad)
    conn.close()
    return {'submission': submission, 'student': student_info, 'answers': answers}

@router.put('/{exam_id}/submissions/{student_id}/approve')
async def approve_submission(request: Request, room_id: int, exam_id: int, student_id: int, body: ApproveSubmissionRequest, user: dict=Depends(get_current_user)):
    """Teacher approves a submission, optionally overriding AI scores."""
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can approve submissions')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail='Unauthorized')
    cursor.execute('SELECT * FROM submissions WHERE exam_id = ? AND student_id = ?', (exam_id, student_id))
    submission = cursor.fetchone()
    if not submission:
        conn.close()
        raise HTTPException(status_code=404, detail='Submission not found')
    submission_id = submission['id']
    total_teacher_score = 0.0
    cursor.execute('SELECT * FROM submission_answers WHERE submission_id = ?', (submission_id,))
    all_answers = cursor.fetchall()
    for ans in all_answers:
        q_id_str = str(ans['question_id'])
        t_score = body.teacher_scores.get(q_id_str) if body.teacher_scores else None
        t_comment = body.teacher_comments.get(q_id_str) if body.teacher_comments else None
        if t_score is not None:
            cursor.execute('UPDATE submission_answers SET teacher_score = ?, teacher_comment = ? WHERE id = ?', (t_score, t_comment, ans['id']))
            total_teacher_score += float(t_score)
        else:
            total_teacher_score += float(ans['ai_score'] or 0)
            if t_comment:
                cursor.execute('UPDATE submission_answers SET teacher_comment = ? WHERE id = ?', (t_comment, ans['id']))
    cursor.execute("UPDATE submissions SET status = 'approved', total_score = ? WHERE id = ?", (round(total_teacher_score, 1), submission_id))
    conn.commit()
    conn.close()
    return {'message': 'Submission approved', 'total_score': round(total_teacher_score, 1)}

@router.post('/{exam_id}/questions/{question_id}/rescore')
async def rescore_question(room_id: int, exam_id: int, question_id: int, user: dict=Depends(get_current_user)):
    """Force AI to rescore a specific question for all submissions in an exam."""
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can rescore')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail='Unauthorized')
    cursor.execute('SELECT submission_id FROM submission_answers WHERE question_id = ?', (question_id,))
    subs = cursor.fetchall()
    for row in subs:
        await grading_queue.put({'submission_id': row['submission_id'], 'room_id': room_id, 'exam_id': exam_id, 'question_id': question_id})
    conn.close()
    return {'message': f'เพิ่มการประมวลผลข้อนี้ใหม่จำนวน {len(subs)} รายการลงในคิวแล้ว'}

@router.post('/{exam_id}/bulk-approve')
async def bulk_approve(room_id: int, exam_id: int, body: BulkApproveRequest, user: dict=Depends(get_current_user)):
    """Teacher bulk-approves multiple students using AI scores as-is."""
    if user['role'] != 'teacher':
        raise HTTPException(status_code=403, detail='Only teachers can approve submissions')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM rooms WHERE id = ? AND owner_id = ?', (room_id, user['id']))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=403, detail='Unauthorized')
    approved = []
    skipped = []
    for sid in body.student_ids:
        cursor.execute('SELECT id, status FROM submissions WHERE exam_id = ? AND student_id = ?', (exam_id, sid))
        sub = cursor.fetchone()
        if not sub:
            skipped.append({'student_id': sid, 'reason': 'not_submitted'})
            continue
        if sub['status'] in ('ready', 'needs_review'):
            cursor.execute("UPDATE submissions SET status = 'approved' WHERE id = ?", (sub['id'],))
            approved.append(sid)
        else:
            skipped.append({'student_id': sid, 'reason': f"status_{sub['status']}"})
    conn.commit()
    conn.close()
    return {'message': f'Approved {len(approved)} submissions', 'approved_student_ids': approved, 'skipped': skipped}


